from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
import json
import time
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import SPORent, CFAAgreement, TransporterAgreement, Approval, MasState, MasStateBranch, MasPartnerDetails, CFAPartnerDetails, TransporterPartnerDetails, ApprovalWorkflow, ApprovalWorkflowHistory, MasDistrict
import logging
import uuid
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from .models import ChatbotQuestion, ChatbotSession, ChatbotMessage, UserMenuAccess
from .forms import (CustomUserCreationForm, SPORentForm, CFAAgreementForm, 
                               TransporterAgreementForm, PartnerDetailsForm, CFAPartnerDetailsForm, 
                               TransporterPartnerDetailsForm, ChatbotQuestionForm,
                               MasDistrictForm, ApprovalForm)
from django.contrib.auth import update_session_auth_hash
from django.http import JsonResponse

# Set up logging
logger = logging.getLogger(__name__)

def send_spo_rent_creation_email(spo_record, created_by=None):
    """
    Send email notification when a new SPO Rent record is created
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'New SPO Rent Record Created - {spo_record.spo_code} ({spo_record.spo_name})'
        
        # Render email templates
        context = {
            'record': spo_record,
            'created_by': created_by or 'System'
        }
        
        html_message = render_to_string('dashboard/emails/banking_professional_template.html', context)
        text_message = render_to_string('dashboard/emails/banking_professional_template.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"SPO Rent creation email sent successfully for record ID: {spo_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send SPO Rent creation email for record ID {spo_record.id}: {str(e)}")
        return False

def send_approval_creation_email(approval_record, created_by=None):
    """
    Send email notification when a new Approval record is created
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'New Approval Record Created - {approval_record.nfa_no} ({approval_record.name})'
        
        # Render email templates
        context = {
            'record': approval_record,
            'created_by': created_by or 'System',
            'is_update': False
        }
        
        html_message = render_to_string('dashboard/emails/approval_creation_template.html', context)
        text_message = render_to_string('dashboard/emails/approval_creation_template.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"Approval creation email sent successfully for record ID: {approval_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send Approval creation email for record ID {approval_record.id}: {str(e)}")
        return False

def send_approval_update_email(approval_record, updated_by=None):
    """
    Send email notification when an Approval record is updated
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'Approval Record Updated - {approval_record.nfa_no} ({approval_record.name})'
        
        # Render email templates
        context = {
            'record': approval_record,
            'updated_by': updated_by or 'System',
            'is_update': True
        }
        
        html_message = render_to_string('dashboard/emails/approval_update_template.html', context)
        text_message = render_to_string('dashboard/emails/approval_update_template.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"Approval update email sent successfully for record ID: {approval_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send Approval update email for record ID {approval_record.id}: {str(e)}")
        return False

def send_partner_joined_email(spo_record, partner, created_by=None):
    """
    Send email notification when a new partner is added to SPO
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'New Partner Added - SPO {spo_record.spo_code} ({partner.name})'
        
        # Get total partners count
        total_partners = MasPartnerDetails.objects.filter(spo=spo_record).count()
        
        # Render email templates
        context = {
            'spo_record': spo_record,
            'partner': partner,
            'total_partners': total_partners,
            'created_by': created_by or 'System',
            'current_date': timezone.now(),
            'dashboard_url': f'{settings.BASE_URL}/dashboard/spo-rent/' if hasattr(settings, 'BASE_URL') else 'http://localhost:8000/dashboard/spo-rent/'
        }
        
        html_message = render_to_string('dashboard/emails/partner_joined_notification.html', context)
        text_message = render_to_string('dashboard/emails/partner_joined_notification.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"Partner joined email sent successfully for SPO ID: {spo_record.id}, Partner: {partner.name}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send partner joined email for SPO ID {spo_record.id}, Partner {partner.name}: {str(e)}")
        return False

def send_cfa_agreement_creation_email(cfa_record, created_by=None):
    """
    Send email notification when a new CFA Agreement record is created
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'New CFA Agreement Record Created - {cfa_record.cfa_code} ({cfa_record.cfa_name})'
        
        # Render email templates
        context = {
            'record': cfa_record,
            'created_by': created_by or 'System',
            'updated_by': created_by or 'System',  # Add this for template compatibility
            'is_update': False  # Add this to indicate it's a creation, not update
        }
        
        html_message = render_to_string('dashboard/emails/cfa_agreement_creation_notification.html', context)
        text_message = render_to_string('dashboard/emails/cfa_agreement_creation_notification.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"CFA Agreement creation email sent successfully for record ID: {cfa_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send CFA Agreement creation email for record ID {cfa_record.id}: {str(e)}")
        return False

def send_cfa_partner_joined_email(cfa_record, partner, created_by=None):
    """
    Send email notification when a new partner is added to CFA Agreement
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'New Partner Added - CFA {cfa_record.cfa_code} ({partner.name})'
        
        # Get total partners count
        total_partners = CFAPartnerDetails.objects.filter(cfa_agreement=cfa_record).count()
        
        # Render email templates
        context = {
            'cfa_record': cfa_record,
            'partner': partner,
            'total_partners': total_partners,
            'created_by': created_by or 'System',
            'current_date': timezone.now(),
            'dashboard_url': f'{settings.BASE_URL}/dashboard/cfa-agreement/' if hasattr(settings, 'BASE_URL') else 'http://localhost:8000/dashboard/cfa-agreement/'
        }
        
        html_message = render_to_string('dashboard/emails/cfa_partner_joined_notification.html', context)
        text_message = render_to_string('dashboard/emails/cfa_partner_joined_notification.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"CFA Partner joined email sent successfully for CFA ID: {cfa_record.id}, Partner: {partner.name}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send CFA partner joined email for CFA ID {cfa_record.id}, Partner {partner.name}: {str(e)}")
        return False

def send_cfa_agreement_update_email(cfa_record, updated_by=None):
    """
    Send email notification when a CFA Agreement record is updated
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'CFA Agreement Record Updated - {cfa_record.cfa_code} ({cfa_record.cfa_name})'
        
        # Render email templates
        context = {
            'record': cfa_record,
            'created_by': updated_by or 'System',  # Add this for template compatibility
            'updated_by': updated_by or 'System',
            'is_update': True
        }
        
        html_message = render_to_string('dashboard/emails/cfa_agreement_creation_notification.html', context)
        text_message = render_to_string('dashboard/emails/cfa_agreement_creation_notification.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"CFA Agreement update email sent successfully for record ID: {cfa_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send CFA Agreement update email for record ID {cfa_record.id}: {str(e)}")
        return False

def send_transporter_agreement_creation_email(transporter_record, created_by=None):
    """
    Send email notification when a new Transporter Agreement record is created
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'New Transporter Agreement Created - {transporter_record.transporter_code} ({transporter_record.transporter_name})'
        
        # Render email templates
        context = {
            'record': transporter_record,
            'created_by': created_by or 'System'
        }
        
        html_message = render_to_string('dashboard/emails/banking_professional_template.html', context)
        text_message = render_to_string('dashboard/emails/banking_professional_template.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"Transporter Agreement creation email sent successfully for record ID: {transporter_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send Transporter Agreement creation email for record ID {transporter_record.id}: {str(e)}")
        return False

def send_transporter_agreement_update_email(transporter_record, updated_by=None):
    """
    Send email notification when a Transporter Agreement is updated
    """
    try:
        # Email recipients
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'Transporter Agreement Updated - {transporter_record.transporter_code} ({transporter_record.transporter_name})'
        
        # Render email templates
        context = {
            'record': transporter_record,
            'created_by': updated_by or 'System',
            'is_update': True
        }
        
        html_message = render_to_string('dashboard/emails/banking_professional_template.html', context)
        text_message = render_to_string('dashboard/emails/banking_professional_template.txt', context)
        
        # Create email message with HTML as primary content
        email = EmailMessage(
            subject=subject,
            body=html_message,  # Use HTML as primary content
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Add plain text as alternative
        email.alternatives = [(text_message, 'text/plain')]
        
        # Send email
        email.send()
        
        logger.info(f"Transporter Agreement update email sent successfully for record ID: {transporter_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send Transporter Agreement update email for record ID {transporter_record.id}: {str(e)}")
        return False

@login_required
def dashboard(request):
    # Get current date for expiry calculations
    today = timezone.now().date()
    
    # Get SPO expiry data
    spo_expired = SPORent.objects.filter(
        rental_to_date__lt=today,
        rental_to_date__isnull=False
    ).order_by('rental_to_date')[:10]  # Limit to 10 most recent expired
    
    # Add expired days count for expired records
    for record in spo_expired:
        record.expired_days_count = (today - record.rental_to_date).days
    
    spo_active = SPORent.objects.filter(
        rental_to_date__gte=today,
        rental_to_date__isnull=False
    ).order_by('rental_to_date')[:10]  # Limit to 10 most recent active
    
    # Get SPO records expiring soon (within 30 days)
    spo_expiring_soon = SPORent.objects.filter(
        rental_to_date__gte=today,
        rental_to_date__lte=today + timezone.timedelta(days=30),
        rental_to_date__isnull=False
    ).order_by('rental_to_date')[:5]  # Limit to 5 most urgent
    
    # Add days until expiry for active records
    for record in spo_active:
        record.days_until_expiry = (record.rental_to_date - today).days
    
    for record in spo_expiring_soon:
        record.days_until_expiry = (record.rental_to_date - today).days
    
    spo_expired_count = SPORent.objects.filter(
        rental_to_date__lt=today,
        rental_to_date__isnull=False
    ).count()
    
    spo_active_count = SPORent.objects.filter(
        rental_to_date__gte=today,
        rental_to_date__isnull=False
    ).count()
    
    # Get CFA expiry data
    cfa_expired = CFAAgreement.objects.filter(
        agreement_to_date__lt=today,
        agreement_to_date__isnull=False
    ).order_by('agreement_to_date')[:10]  # Limit to 10 most recent expired
    
    # Add expired days count for expired records
    for record in cfa_expired:
        record.expired_days_count = (today - record.agreement_to_date).days
    
    cfa_active = CFAAgreement.objects.filter(
        agreement_to_date__gte=today,
        agreement_to_date__isnull=False
    ).order_by('agreement_to_date')[:10]  # Limit to 10 most recent active
    
    # Add days until expiry for active records
    for record in cfa_active:
        record.days_until_expiry = (record.agreement_to_date - today).days
    
    cfa_expired_count = CFAAgreement.objects.filter(
        agreement_to_date__lt=today,
        agreement_to_date__isnull=False
    ).count()
    
    cfa_active_count = CFAAgreement.objects.filter(
        agreement_to_date__gte=today,
        agreement_to_date__isnull=False
    ).count()
    
    # Get Transporter expiry data
    transporter_expired = TransporterAgreement.objects.filter(
        agreement_to_date__lt=today,
        agreement_to_date__isnull=False
    ).order_by('agreement_to_date')[:10]  # Limit to 10 most recent expired
    
    # Add expired days count for expired records
    for record in transporter_expired:
        record.expired_days_count = (today - record.agreement_to_date).days
    
    transporter_active = TransporterAgreement.objects.filter(
        agreement_to_date__gte=today,
        agreement_to_date__isnull=False
    ).order_by('agreement_to_date')[:10]  # Limit to 10 most recent active
    
    # Add days until expiry for active records
    for record in transporter_active:
        record.days_until_expiry = (record.agreement_to_date - today).days
    
    transporter_expired_count = TransporterAgreement.objects.filter(
        agreement_to_date__lt=today,
        agreement_to_date__isnull=False
    ).count()
    
    transporter_active_count = TransporterAgreement.objects.filter(
        agreement_to_date__gte=today,
        agreement_to_date__isnull=False
    ).count()
    
    # Get statistics for dashboard cards
    context = {
        'spo_rent_count': SPORent.objects.count(),
        'cfa_agreement_count': CFAAgreement.objects.count(),
        'transporter_agreement_count': TransporterAgreement.objects.count(),
        'user_count': User.objects.count(),
        
        # SPO expiry data
        'spo_expired': spo_expired,
        'spo_active': spo_active,
        'spo_expiring_soon': spo_expiring_soon,
        'spo_expired_count': spo_expired_count,
        'spo_active_count': spo_active_count,
        
        # CFA expiry data
        'cfa_expired': cfa_expired,
        'cfa_active': cfa_active,
        'cfa_expired_count': cfa_expired_count,
        'cfa_active_count': cfa_active_count,
        
        # Transporter expiry data
        'transporter_expired': transporter_expired,
        'transporter_active': transporter_active,
        'transporter_expired_count': transporter_expired_count,
        'transporter_active_count': transporter_active_count,
        
        'today': today,
    }
    return render(request, 'dashboard/dashboard.html', context)

@login_required
def user_list(request):
    users = User.objects.all()
    return render(request, 'dashboard/user_list.html', {'users': users})

@login_required
def add_user(request):
    """Add a new user"""
    if request.method == 'POST':
        name = request.POST.get('name')
        username = request.POST.get('username')
        mobile_no = request.POST.get('mobile_no')
        email = request.POST.get('email')
        gender = request.POST.get('gender')
        password = request.POST.get('password')
        is_staff = request.POST.get('is_staff') == 'on'
        is_superuser = request.POST.get('is_superuser') == 'on'
        is_admin = request.POST.get('is_admin') == 'on'
        
        if username and email and password and name and mobile_no and gender:
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Employee Code already exists.')
            elif User.objects.filter(email=email).exists():
                messages.error(request, 'Email already exists.')
            else:
                # Split name into first and last name
                name_parts = name.strip().split(' ', 1)
                first_name = name_parts[0] if name_parts else ''
                last_name = name_parts[1] if len(name_parts) > 1 else ''
                
                # Create the user with basic permissions
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    password=password,
                    is_staff=is_staff,
                    is_superuser=is_superuser
                )
                
                # Handle admin permissions
                if is_admin:
                    user.is_staff = True
                    user.is_superuser = True
                    user.save()
                
                # Store additional user details in user profile or custom fields
                # You can extend the User model or create a UserProfile model for mobile_no and gender
                # For now, we'll store them in the first_name and last_name fields temporarily
                # In a production system, you'd want to create proper models for these fields
                
                messages.success(request, f'User {name} (Emp.Code: {username}) created successfully.')
                return redirect('user_list')
        else:
            messages.error(request, 'Please fill in all required fields.')
    
    return render(request, 'dashboard/add_user.html')

@login_required
def edit_user(request, user_id):
    """Edit an existing user"""
    try:
        user = User.objects.get(id=user_id)
        if request.method == 'POST':
            user.username = request.POST.get('username')
            user.email = request.POST.get('email')
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.is_staff = request.POST.get('is_staff') == 'on'
            user.is_superuser = request.POST.get('is_superuser') == 'on'
            user.is_admin = request.POST.get('is_admin') == 'on'
            user.is_active = request.POST.get('is_active') == 'on'
            
            if user.username and user.email:
                # Handle admin permissions
                if user.is_admin:
                    user.is_staff = True
                    user.is_superuser = True
                
                user.save()
                messages.success(request, f'User {user.username} updated successfully.')
                return redirect('user_list')
            else:
                messages.error(request, 'Please fill in all required fields.')
        
        return render(request, 'dashboard/edit_user.html', {'user': user})
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('user_list')

@login_required
def delete_user(request, user_id):
    """Delete a user"""
    try:
        user = User.objects.get(id=user_id)
        if request.user.id == user.id:
            messages.error(request, 'You cannot delete your own account.')
        else:
            username = user.username
            user.delete()
            messages.success(request, f'User {username} deleted successfully.')
        return redirect('user_list')
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('user_list')

@login_required
def view_user(request, user_id):
    """View user details"""
    try:
        user = User.objects.get(id=user_id)
        return render(request, 'dashboard/view_user.html', {'user': user})
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('user_list')

@login_required
def reset_user_password(request, user_id):
    """Reset password for a specific user"""
    try:
        user = User.objects.get(id=user_id)
        if request.method == 'POST':
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if new_password and confirm_password:
                if new_password == confirm_password:
                    if len(new_password) >= 8:
                        user.set_password(new_password)
                        user.save()
                        messages.success(request, f'Password for user {user.username} has been reset successfully.')
                        return redirect('user_list')
                    else:
                        messages.error(request, 'Password must be at least 8 characters long.')
                else:
                    messages.error(request, 'Passwords do not match.')
            else:
                messages.error(request, 'Please fill in all fields.')
        
        return render(request, 'dashboard/reset_password.html', {'user': user})
    except User.DoesNotExist:
        messages.error(request, 'User not found.')
        return redirect('user_list')

@login_required
def change_password(request):
    """Change password for the current user"""
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if current_password and new_password and confirm_password:
            if request.user.check_password(current_password):
                if new_password == confirm_password:
                    if len(new_password) >= 8:
                        request.user.set_password(new_password)
                        request.user.save()
                        messages.success(request, 'Your password has been changed successfully.')
                        return redirect('dashboard')
                    else:
                        messages.error(request, 'New password must be at least 8 characters long.')
                else:
                    messages.error(request, 'New passwords do not match.')
            else:
                messages.error(request, 'Current password is incorrect.')
        else:
            messages.error(request, 'Please fill in all fields.')
    
    return render(request, 'dashboard/change_password.html')

# SPO Rent Views
@login_required
def spo_rent_list(request):
    # Force fresh database query - no caching
    from django.db import connection
    connection.close()
    
    # Get fresh data from database ONLY - no static data
    records = SPORent.objects.select_related('state', 'branch').all().order_by('-created_at')
    
    # Apply filters if provided
    spo_name = request.GET.get('spo_name', '').strip()
    spo_code = request.GET.get('spo_code', '').strip()
    branch_name = request.GET.get('branch_name', '').strip()
    district_name = request.GET.get('district_name', '').strip()
    state = request.GET.get('state', '').strip()
    owner_name = request.GET.get('owner_name', '').strip()
    rent_status = request.GET.get('rent_status', '').strip()
    
    if spo_name:
        records = records.filter(spo_name__icontains=spo_name)
    if spo_code:
        records = records.filter(spo_code__icontains=spo_code)
    if branch_name:
        records = records.filter(branch__state_branch_name__icontains=branch_name)
    if district_name:
        records = records.filter(district_code__icontains=district_name)
    if state:
        records = records.filter(state__state_name__icontains=state)
    if owner_name:
        records = records.filter(owner_name__icontains=owner_name)
    if rent_status:
        records = records.filter(status__iexact=rent_status)
    
    # Handle rent amount range filtering
    rent_amount = request.GET.get('rent_amount', '').strip()
    if rent_amount:
        if rent_amount == '0-5000':
            records = records.filter(security_deposit_paid__gte=0, security_deposit_paid__lte=5000)
        elif rent_amount == '5000-10000':
            records = records.filter(security_deposit_paid__gte=5000, security_deposit_paid__lte=10000)
        elif rent_amount == '10000-20000':
            records = records.filter(security_deposit_paid__gte=10000, security_deposit_paid__lte=20000)
        elif rent_amount == '20000+':
            records = records.filter(security_deposit_paid__gte=20000)
    
    # Get exact count from database
    total_records = records.count()
    
    # Add cache-busting timestamp
    import time
    cache_buster = int(time.time())
    
    # Add debug information
    debug_info = {
        'database_records': total_records,
        'cache_buster': cache_buster,
        'timestamp': time.strftime('%Y-%m-%d %H:%M:%S'),
        'is_empty': total_records == 0
    }
    
    # Get states and branches for export filters
    from .models import MasState, MasStateBranch
    states = MasState.objects.all().order_by('state_name')
    branches = MasStateBranch.objects.all().order_by('state_branch_name')
    
    return render(request, 'dashboard/spo_rent_list.html', {
        'records': records, 
        'cache_buster': cache_buster,
        'total_records': total_records,
        'debug_info': debug_info,
        'is_empty': total_records == 0,
        'states': states,
        'branches': branches
    })

@login_required
def spo_rent_export_excel(request):
    """Export SPO Rent records to Excel format with enhanced filtering"""
    from .models import MasState, MasStateBranch
    
    # Get export options
    include_summary = request.GET.get('include_summary', 'true').lower() == 'true'
    include_partners = request.GET.get('include_partners', 'true').lower() == 'true'
    
    # Get filtered records
    records = SPORent.objects.select_related('state', 'branch').all()
    
    # Apply filters if provided
    # SPO State Filter
    state_id = request.GET.get('state', '').strip()
    if state_id:
        records = records.filter(state_id=state_id)
    
    # Date Range Filters
    from_date = request.GET.get('from_date', '').strip()
    to_date = request.GET.get('to_date', '').strip()
    
    if from_date:
        records = records.filter(created_at__gte=from_date)
    if to_date:
        records = records.filter(created_at__lte=to_date)
    
    # Status Filter
    status = request.GET.get('status', '').strip()
    if status:
        records = records.filter(status__iexact=status)
    
    # Additional Filters
    spo_code = request.GET.get('spo_code', '').strip()
    spo_name = request.GET.get('spo_name', '').strip()
    branch_id = request.GET.get('branch', '').strip()
    owner_name = request.GET.get('owner_name', '').strip()
    rent_amount_min = request.GET.get('rent_amount_min', '').strip()
    rent_amount_max = request.GET.get('rent_amount_max', '').strip()
    
    if spo_code:
        records = records.filter(spo_code__icontains=spo_code)
    if spo_name:
        records = records.filter(spo_name__icontains=spo_name)
    if branch_id:
        records = records.filter(branch_id=branch_id)
    if owner_name:
        records = records.filter(owner_name__icontains=owner_name)
    if rent_amount_min:
        records = records.filter(rent_pm__gte=float(rent_amount_min))
    if rent_amount_max:
        records = records.filter(rent_pm__lte=float(rent_amount_max))
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "SPO Rent Records"
    
    # Add summary information if requested
    current_row = 1
    if include_summary:
        ws[f'A{current_row}'] = "SPO Rent Records Export"
        current_row += 1
        ws[f'A{current_row}'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        current_row += 1
        ws[f'A{current_row}'] = f"Total Records: {records.count()}"
        current_row += 1
        ws[f'A{current_row}'] = f"Include Partners: {include_partners}"
        current_row += 1
    
    # Show active filters
    active_filters = []
    if state_id: 
        state_name = MasState.objects.get(id=state_id).state_name if MasState.objects.filter(id=state_id).exists() else state_id
        active_filters.append(f"State: {state_name}")
    if from_date: active_filters.append(f"From Date: {from_date}")
    if to_date: active_filters.append(f"To Date: {to_date}")
    if status: active_filters.append(f"Status: {status}")
    if spo_code: active_filters.append(f"SPO Code: {spo_code}")
    if spo_name: active_filters.append(f"SPO Name: {spo_name}")
    if branch_id: 
        branch_name = MasStateBranch.objects.get(id=branch_id).state_branch_name if MasStateBranch.objects.filter(id=branch_id).exists() else branch_id
        active_filters.append(f"Branch: {branch_name}")
    if owner_name: active_filters.append(f"Owner Name: {owner_name}")
    if rent_amount_min: active_filters.append(f"Min Rent: ₹{rent_amount_min}")
    if rent_amount_max: active_filters.append(f"Max Rent: ₹{rent_amount_max}")
    
    if include_summary:
        ws[f'A{current_row}'] = f"Filters Applied: {'Yes' if active_filters else 'No'}"
        current_row += 1
        if active_filters:
            ws[f'A{current_row}'] = "Active Filters:"
            current_row += 1
            for filter_text in active_filters:
                ws[f'A{current_row}'] = f"  • {filter_text}"
                current_row += 1
        
        # Style summary section
        summary_font = Font(bold=True, size=12)
        summary_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
        
        # Apply styling to summary rows
        for row in range(1, current_row):
            cell = ws.cell(row=row, column=1)
            cell.font = summary_font
            cell.fill = summary_fill
        
        # Start data from row after summary + 1 blank row
        data_start_row = current_row + 1
    else:
        # No summary, start data from first row
        data_start_row = 1
    
    # Define headers
    headers = [
        'S.No', 'SPO Code', 'SPO Name', 'State', 'State Code', 'Branch', 'Branch Code', 
        'District Code', 'Structure Group', 'SPO Status', 'Inception Date', 'Renewal With',
        'Godown Address', 'Owner Name', 'Owner Code', 'Owner Email', 'Owner Phone', 'Owner Address', 
        'Owner GST', 'Owner PAN', 'Pin Code', 'Nature of Construction', 'Sale Organization',
        'Stamp No', 'Stamp Name', 'Partner Type', 'Bank Account Name', 'Bank Account No', 'Bank Name',
        'Bank Branch', 'Bank IFSC', 'Destination Code', 'Office Sq.Ft', 'Open Space Sq.Ft',
        'Total Space', 'Capacity', 'Rent From Date', 'Rent To Date', 'Days Count',
        'Security Deposit', 'Security Deposit Doc', 'Monthly Rent', 'Yearly Hike %',
        'Latitude', 'Longitude', 'Vacation Letter', 'Remarks', 'Status'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    for row_idx, record in enumerate(records, data_start_row + 1):
        # Get partner details if requested
        partner_details = ""
        if include_partners:
            from .models import MasPartnerDetails
            partners = MasPartnerDetails.objects.filter(spo_id=record.id)
            if partners.exists():
                partner_names = [f"{p.name} ({p.gender}, {p.age} years)" for p in partners]
                partner_details = "; ".join(partner_names)
            else:
                partner_details = "No partners"
        
        # Add partner details column to headers if not already present
        if include_partners and 'Partner Details' not in headers:
            headers.append('Partner Details')
            # Add header for partner details column
            partner_col = len(headers)
            cell = ws.cell(row=data_start_row, column=partner_col, value='Partner Details')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        row_data = [
            row_idx - data_start_row,  # S.No
            record.spo_code or 'N/A',
            record.spo_name or 'N/A',
            record.state.state_name if record.state else 'N/A',
            record.state.state_code if record.state else 'N/A',
            record.branch.state_branch_name if record.branch else 'N/A',
            record.branch.state_branch_code if record.branch else 'N/A',
            record.district_code or 'N/A',
            record.stru_grp or 'N/A',
            record.cfa_status or 'N/A',
            record.inception_date.strftime('%Y-%m-%d') if record.inception_date else 'N/A',
            record.renewal_with or 'N/A',
            record.godown_address or 'N/A',
            record.owner_name or 'N/A',
            record.owner_code or 'N/A',
            record.cfa_mail_id or 'N/A',
            record.owner_contact_no or 'N/A',
            record.owner_address or 'N/A',
            record.owner_gst or 'N/A',
            record.owner_pan or 'N/A',
            record.pin_code or 'N/A',
            record.nature_of_construction.name if record.nature_of_construction else 'N/A',
            record.sale_organization or 'N/A',
            record.stamp_no or 'N/A',
            record.stamp_name or 'N/A',
            record.partner_type.name if record.partner_type else 'N/A',
            record.bank_account_name or 'N/A',
            record.bank_account_no or 'N/A',
            record.bank_name or 'N/A',
            record.bank_branch_name or 'N/A',
            record.bank_ifsc_code or 'N/A',
            record.destination_code or 'N/A',
            record.office_sqft or 'N/A',
            record.open_space_sqft or 'N/A',
            record.total_space or 'N/A',
            record.capacity or 'N/A',
            record.rental_from_date.strftime('%Y-%m-%d') if record.rental_from_date else 'N/A',
            record.rental_to_date.strftime('%Y-%m-%d') if record.rental_to_date else 'N/A',
            record.days_count or 'N/A',
            f"₹{record.security_deposit_paid:.2f}" if record.security_deposit_paid else 'N/A',
            record.security_deposit_doc or 'N/A',
            f"₹{record.rent_pm:.2f}" if record.rent_pm else 'N/A',
            f"{record.yearly_hike_percent:.2f}%" if record.yearly_hike_percent else 'N/A',
            f"{record.latitude:.6f}" if record.latitude else 'N/A',
            f"{record.longitude:.6f}" if record.longitude else 'N/A',
            'Yes' if record.vacation_letter else 'No',
            record.remarks or 'N/A',
            record.status or 'N/A'
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="SPO_Rent_Records_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save to response
    wb.save(response)
    
    return response

@login_required
def spo_rent_create(request):
    """Create a new SPO Rent record - using standard Django form approach like Approval page"""
    if request.method == 'POST':
        form = SPORentForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Handle branch selection and auto-fill district code
                branch_id = request.POST.get('branch')
                if branch_id:
                    try:
                        branch = MasStateBranch.objects.get(id=branch_id)
                        form.instance.branch = branch
                        # Auto-fill district code from branch
                        if not form.instance.district_code:
                            form.instance.district_code = branch.state_branch_code
                    except MasStateBranch.DoesNotExist:
                        messages.warning(request, f'Selected branch (ID: {branch_id}) not found. Using form data.')
                
                # Handle state selection
                state_id = request.POST.get('state')
                if state_id:
                    try:
                        state = MasState.objects.get(id=state_id)
                        form.instance.state = state
                    except MasState.DoesNotExist:
                        messages.warning(request, f'Selected state (ID: {state_id}) not found. Using form data.')
                
                # Save the form
                spo_record = form.save()
                
                # Create approval workflow for the new record
                try:
                    workflow = create_approval_workflow(
                        record_type='spo_rent',
                        record_id=spo_record.id,
                        record_code=spo_record.spo_code,
                        submitted_by=request.user
                    )
                    if workflow:
                        logger.info(f"Approval workflow created for SPO Rent record: {spo_record.spo_code}")
                    else:
                        logger.warning(f"Failed to create approval workflow for SPO Rent record: {spo_record.spo_code}")
                except Exception as workflow_error:
                    logger.error(f"Error creating approval workflow: {str(workflow_error)}")
                
                # Send email notification for new SPO Rent record creation
                try:
                    created_by = request.user.get_full_name() or request.user.username
                    email_sent = send_spo_rent_creation_email(spo_record, created_by)
                    if email_sent:
                        logger.info(f"Email notification sent for new SPO Rent record: {spo_record.spo_code}")
                    else:
                        logger.warning(f"Failed to send email notification for new SPO Rent record: {spo_record.spo_code}")
                except Exception as email_error:
                    logger.error(f"Error sending email notification: {str(email_error)}")
                
                # Check if this is an AJAX request (from modal)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'SPO Rent record "{spo_record.spo_name}" created successfully!',
                        'record_id': spo_record.id
                    })
                else:
                    messages.success(request, f'SPO Rent record "{spo_record.spo_name}" created successfully!')
                    return redirect('spo_rent_list')
                
            except Exception as e:
                error_msg = f'Error creating SPO Rent record: {str(e)}'
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': error_msg
                    })
                else:
                    messages.error(request, error_msg)
                    print(f"Error creating SPO Rent record: {e}")
        else:
            # Show form errors
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            
            error_msg = f'Please correct the following errors: {" ".join(error_messages)}' if error_messages else 'Please correct the errors below.'
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': error_msg,
                    'errors': form.errors
                })
            else:
                if error_messages:
                    messages.error(request, error_msg)
                else:
                    messages.error(request, 'Please correct the errors below.')
    else:
        form = SPORentForm()
    
    return render(request, 'dashboard/spo_rent_form.html', {
        'form': form, 
        'title': 'Add SPO Rent', 
        'is_edit': False
    })

@login_required
def spo_rent_edit(request, record_id):
    """Edit an existing SPO Rent record - using standard Django form approach like Approval page"""
    record = get_object_or_404(SPORent, id=record_id)
    
    if request.method == 'POST':
        form = SPORentForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            try:
                # Handle branch selection and auto-fill district code
                branch_id = request.POST.get('branch')
                if branch_id:
                    try:
                        branch = MasStateBranch.objects.get(id=branch_id)
                        form.instance.branch = branch
                        # Auto-fill district code from branch
                        if not form.instance.district_code:
                            form.instance.district_code = branch.state_branch_code
                    except MasStateBranch.DoesNotExist:
                        messages.warning(request, f'Selected branch (ID: {branch_id}) not found. Using form data.')
                
                # Handle state selection
                state_id = request.POST.get('state')
                if state_id:
                    try:
                        state = MasState.objects.get(id=state_id)
                        form.instance.state = state
                    except MasState.DoesNotExist:
                        messages.warning(request, f'Selected state (ID: {state_id}) not found. Using form data.')
                
                form.save()
                messages.success(request, 'SPO Rent record updated successfully!')
                return redirect('spo_rent_list')
            except Exception as e:
                messages.error(request, f'Error updating SPO Rent record: {str(e)}')
                print(f"Error updating SPO Rent record: {e}")
        else:
            # Show form errors
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            
            if error_messages:
                messages.error(request, f'Please correct the following errors: {" ".join(error_messages)}')
            else:
                messages.error(request, 'Please correct the errors below.')
    else:
        form = SPORentForm(instance=record)
    
    return render(request, 'dashboard/spo_rent_form.html', {
        'form': form, 
        'title': 'Edit SPO Rent', 
        'record': record, 
        'is_edit': True
    })

@login_required
def spo_rent_view(request, record_id):
    record = get_object_or_404(SPORent, id=record_id)
    # Get all partners associated with this SPO
    partners = MasPartnerDetails.objects.filter(spo=record).order_by('created_at')
    return render(request, 'dashboard/spo_rent_view.html', {
        'record': record,
        'partners': partners
    })

@login_required
def spo_rent_delete(request, record_id):
    record = get_object_or_404(SPORent, id=record_id)
    if request.method == 'POST':
        record.delete()
        messages.success(request, 'SPO Rent record deleted successfully!')
        return redirect('spo_rent_list')
    return render(request, 'dashboard/spo_rent_delete.html', {'record': record})

@login_required
def spo_rent_pdf(request, record_id):
    record = get_object_or_404(SPORent, id=record_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="SPO_Rent_{record.spo_code}_{record.spo_name}.pdf"'
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30, alignment=TA_CENTER, textColor=colors.darkblue
    )
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=12, spaceBefore=20, textColor=colors.darkblue
    )
    normal_style = styles['Normal']

    elements.append(Paragraph(f"SPO Rent Details", title_style))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>SPO Code:</b> {record.spo_code}", normal_style))
    elements.append(Paragraph(f"<b>SPO Name:</b> {record.spo_name}", normal_style))
    elements.append(Paragraph(f"<b>Generated On:</b> {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", normal_style))
    elements.append(Spacer(1, 20))

    # Basic Information Section
    elements.append(Paragraph("Basic Information", heading_style))
    basic_data = [
        ['Field', 'Value'],
        ['SPO Code', record.spo_code or 'Not specified'],
        ['SPO Name', record.spo_name or 'Not specified'],
        ['Structure Group', record.stru_grp or 'Not specified'],
        ['SPO Status', record.cfa_status or 'Not specified'],
        ['Inception Date', record.inception_date.strftime('%B %d, %Y') if record.inception_date else 'Not specified'],
        ['Status', record.status or 'Not specified'],
    ]
    basic_table = Table(basic_data, colWidths=[2*inch, 4*inch])
    basic_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(basic_table)
    elements.append(Spacer(1, 20))

    # Location Information Section
    elements.append(Paragraph("Location Information", heading_style))
    location_data = [
        ['Field', 'Value'],
        ['State', record.state.state_name if record.state else 'Not specified'],
        ['District Code', record.district_code or 'Not specified'],
        ['Godown Address', record.godown_address or 'Not specified'],
        ['Destination Code', record.destination_code or 'Not specified'],
        ['Office Sq.Ft', record.office_sqft or 'Not specified'],
        ['Open Space Sq.Ft', record.open_space_sqft or 'Not specified'],
        ['Total Space', record.total_space or 'Not specified'],
        ['Capacity', record.capacity or 'Not specified'],
    ]
    location_table = Table(location_data, colWidths=[2*inch, 4*inch])
    location_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(location_table)
    elements.append(Spacer(1, 20))

    # Owner Information Section
    elements.append(Paragraph("Owner Information", heading_style))
    owner_data = [
        ['Field', 'Value'],
        ['Owner Name', record.owner_name or 'Not specified'],
        ['Owner Contact No', record.owner_contact_no or 'Not specified'],
        ['Owner Code', record.owner_code or 'Not specified'],
        ['Owner Address', record.owner_address or 'Not specified'],
        ['Owner GST', record.owner_gst or 'Not specified'],
        ['Owner PAN', record.owner_pan or 'Not specified'],
        
    ]
    owner_table = Table(owner_data, colWidths=[2*inch, 4*inch])
    owner_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(owner_table)
    elements.append(Spacer(1, 20))

    # Bank Information Section
    elements.append(Paragraph("Bank Information", heading_style))
    bank_data = [
        ['Field', 'Value'],
        ['Bank Account Name', record.bank_account_name or 'Not specified'],
        ['Bank Account No', record.bank_account_no or 'Not specified'],
        ['Bank Name', record.bank_name or 'Not specified'],
        ['Bank Branch Name', record.bank_branch_name or 'Not specified'],
        ['Bank IFSC Code', record.bank_ifsc_code or 'Not specified'],
    ]
    bank_table = Table(bank_data, colWidths=[2*inch, 4*inch])
    bank_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(bank_table)
    elements.append(Spacer(1, 20))

    # Rental Information Section
    elements.append(Paragraph("Rental Information", heading_style))
    rental_data = [
        ['Field', 'Value'],
        ['Rental From Date', record.rental_from_date.strftime('%B %d, %Y') if record.rental_from_date else 'Not specified'],
        ['Rental To Date', record.rental_to_date.strftime('%B %d, %Y') if record.rental_to_date else 'Not specified'],
        ['Days Count', record.days_count or 'Not specified'],
        ['Rent PM', f"₹{record.rent_pm}" if record.rent_pm else 'Not specified'],
        ['Yearly Hike %', f"{record.yearly_hike_percent}%" if record.yearly_hike_percent else 'Not specified'],
        ['Security Deposit Paid', f"₹{record.security_deposit_paid}" if record.security_deposit_paid else 'Not specified'],
        ['Security Deposit Doc', record.security_deposit_doc or 'Not specified'],
    ]
    rental_table = Table(rental_data, colWidths=[2*inch, 4*inch])
    rental_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(rental_table)
    elements.append(Spacer(1, 20))

    # Additional Information Section
    elements.append(Paragraph("Additional Information", heading_style))
    additional_data = [
        ['Field', 'Value'],
        ['Latitude', record.latitude or 'Not specified'],
        ['Longitude', record.longitude or 'Not specified'],
        ['Created', record.created_at.strftime('%B %d, %Y at %I:%M %p') if record.created_at else 'Not specified'],
        ['Last Updated', record.updated_at.strftime('%B %d, %Y at %I:%M %p') if record.updated_at else 'Not specified'],
        ['Remarks', record.remarks or 'No remarks'],
    ]
    additional_table = Table(additional_data, colWidths=[2*inch, 4*inch])
    additional_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(additional_table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

# CFA Agreement Views
@login_required
def cfa_agreement_list(request):
    """List all CFA agreements with advanced filtering"""
    # Get filter parameters
    cfa_name = request.GET.get('cfa_name', '')
    cfa_code = request.GET.get('cfa_code', '')
    branch_name = request.GET.get('branch_name', '')
    district_name = request.GET.get('district_name', '')
    state = request.GET.get('state', '')
    owner_name = request.GET.get('owner_name', '')
    agreement_status = request.GET.get('agreement_status', '')
    agreement_renewal = request.GET.get('agreement_renewal', '')
    
    # Date range filters
    agreement_from_date_min = request.GET.get('agreement_from_date_min', '')
    agreement_from_date_max = request.GET.get('agreement_from_date_max', '')
    agreement_to_date_min = request.GET.get('agreement_to_date_min', '')
    agreement_to_date_max = request.GET.get('agreement_to_date_max', '')
    
    # Security deposit range filters
    security_deposit_min = request.GET.get('security_deposit_min', '')
    security_deposit_max = request.GET.get('security_deposit_max', '')
    
    # Start with all CFA agreements
    records = CFAAgreement.objects.select_related('state').all()
    
    # Apply filters
    if cfa_name:
        records = records.filter(cfa_name__icontains=cfa_name)
    
    if cfa_code:
        records = records.filter(cfa_code__icontains=cfa_code)
    
    if branch_name:
        records = records.filter(branch__state_branch_name__icontains=branch_name)
    
    if district_name:
        records = records.filter(district__name__icontains=district_name)
    
    if state:
        records = records.filter(state__state_name__icontains=state)
    
    if owner_name:
        records = records.filter(owner_name__icontains=owner_name)
    
    if agreement_status:
        records = records.filter(status__iexact=agreement_status)
    
    if agreement_renewal:
        records = records.filter(agreement_renewal__iexact=agreement_renewal)
    
    # Date range filters
    if agreement_from_date_min:
        try:
            from_date = datetime.strptime(agreement_from_date_min, '%Y-%m-%d').date()
            records = records.filter(agreement_from_date__gte=from_date)
        except ValueError:
            pass
    
    if agreement_from_date_max:
        try:
            to_date = datetime.strptime(agreement_from_date_max, '%Y-%m-%d').date()
            records = records.filter(agreement_from_date__lte=to_date)
        except ValueError:
            pass
    
    if agreement_to_date_min:
        try:
            from_date = datetime.strptime(agreement_to_date_min, '%Y-%m-%d').date()
            records = records.filter(agreement_to_date__gte=from_date)
        except ValueError:
            pass
    
    if agreement_to_date_max:
        try:
            to_date = datetime.strptime(agreement_to_date_max, '%Y-%m-%d').date()
            records = records.filter(agreement_to_date__lte=to_date)
        except ValueError:
            pass
    
    # Security deposit range filters
    if security_deposit_min:
        try:
            min_amount = float(security_deposit_min)
            records = records.filter(security_deposit_rs__gte=min_amount)
        except ValueError:
            pass
    
    if security_deposit_max:
        try:
            max_amount = float(security_deposit_max)
            records = records.filter(security_deposit_rs__lte=max_amount)
        except ValueError:
            pass
    
    # Order by CFA code
    records = records.order_by('cfa_code')
    
    # Get states for filter dropdown
    states = MasState.objects.all().order_by('state_name')
    
    # Calculate statistics
    total_records = records.count()
    active_records = records.filter(cfa_status__iexact='Active').count()
    inactive_records = records.filter(cfa_status__iexact='Inactive').count()
    under_notice_records = records.filter(cfa_status__iexact='Under Notice').count()
    closed_records = records.filter(cfa_status__iexact='Closed').count()
    filtered_records = total_records
    
    return render(request, 'dashboard/cfa_agreement_list.html', {
        'records': records,
        'total_records': total_records,
        'active_records': active_records,
        'inactive_records': inactive_records,
        'under_notice_records': under_notice_records,
        'closed_records': closed_records,
        'filtered_records': filtered_records,
        'states': states,
    })

@login_required
def cfa_agreement_export_excel(request):
    """Export CFA Agreement records to Excel format with enhanced filtering"""
    # Get export options
    include_summary = request.GET.get('include_summary', 'true').lower() == 'true'
    include_partners = request.GET.get('include_partners', 'true').lower() == 'true'
    
    # Get filtered records (enhanced filter parameters)
    cfa_name = request.GET.get('cfa_name', '')
    cfa_code = request.GET.get('cfa_code', '')
    branch_name = request.GET.get('branch_name', '')
    district_name = request.GET.get('district_name', '')
    state = request.GET.get('state', '')
    owner_name = request.GET.get('owner_name', '')
    status = request.GET.get('status', '')
    agreement_renewal = request.GET.get('agreement_renewal', '')
    
    # Date range filters
    agreement_from_date_min = request.GET.get('agreement_from_date_min', '')
    agreement_from_date_max = request.GET.get('agreement_from_date_max', '')
    agreement_to_date_min = request.GET.get('agreement_to_date_min', '')
    agreement_to_date_max = request.GET.get('agreement_to_date_max', '')
    
    # Security deposit range filters
    security_deposit_min = request.GET.get('security_deposit_min', '')
    security_deposit_max = request.GET.get('security_deposit_max', '')
    
    # Start with all CFA agreements
    records = CFAAgreement.objects.select_related('state').all()
    
    # Apply filters
    if cfa_name:
        records = records.filter(cfa_name__icontains=cfa_name)
    
    if cfa_code:
        records = records.filter(cfa_code__icontains=cfa_code)
    
    if branch_name:
        records = records.filter(branch__state_branch_name__icontains=branch_name)
    
    if district_name:
        records = records.filter(district__name__icontains=district_name)
    
    if state:
        records = records.filter(state__state_name__icontains=state)
    
    if owner_name:
        records = records.filter(owner_name__icontains=owner_name)
    
    if status:
        records = records.filter(cfa_status__iexact=status)
    
    if agreement_renewal:
        records = records.filter(agreement_renewal__iexact=agreement_renewal)
    
    # Date range filters
    if agreement_from_date_min:
        try:
            from_date = datetime.strptime(agreement_from_date_min, '%Y-%m-%d').date()
            records = records.filter(agreement_from_date__gte=from_date)
        except ValueError:
            pass
    
    if agreement_from_date_max:
        try:
            to_date = datetime.strptime(agreement_from_date_max, '%Y-%m-%d').date()
            records = records.filter(agreement_from_date__lte=to_date)
        except ValueError:
            pass
    
    if agreement_to_date_min:
        try:
            from_date = datetime.strptime(agreement_to_date_min, '%Y-%m-%d').date()
            records = records.filter(agreement_to_date__gte=from_date)
        except ValueError:
            pass
    
    if agreement_to_date_max:
        try:
            to_date = datetime.strptime(agreement_to_date_max, '%Y-%m-%d').date()
            records = records.filter(agreement_to_date__lte=to_date)
        except ValueError:
            pass
    
    # Security deposit range filters
    if security_deposit_min:
        try:
            min_amount = float(security_deposit_min)
            records = records.filter(security_deposit_rs__gte=min_amount)
        except ValueError:
            pass
    
    if security_deposit_max:
        try:
            max_amount = float(security_deposit_max)
            records = records.filter(security_deposit_rs__lte=max_amount)
        except ValueError:
            pass
    
    # Order by CFA code
    records = records.order_by('cfa_code')
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CFA Agreement Records"
    
    # Add summary information
    ws['A1'] = "CFA Agreement Records Export"
    ws['A2'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Total Records: {records.count()}"
    
    # Show active filters
    active_filters = []
    if cfa_name: active_filters.append(f"CFA Name: {cfa_name}")
    if cfa_code: active_filters.append(f"CFA Code: {cfa_code}")
    if branch_name: active_filters.append(f"Branch: {branch_name}")
    if district_name: active_filters.append(f"District: {district_name}")
    if state: active_filters.append(f"State: {state}")
    if owner_name: active_filters.append(f"Owner: {owner_name}")
    if status: active_filters.append(f"Status: {status}")
    if agreement_renewal: active_filters.append(f"Agreement Type: {agreement_renewal}")
    if agreement_from_date_min: active_filters.append(f"From Date (Min): {agreement_from_date_min}")
    if agreement_from_date_max: active_filters.append(f"From Date (Max): {agreement_from_date_max}")
    if agreement_to_date_min: active_filters.append(f"To Date (Min): {agreement_to_date_min}")
    if agreement_to_date_max: active_filters.append(f"To Date (Max): {agreement_to_date_max}")
    if security_deposit_min: active_filters.append(f"Security Deposit (Min): ₹{security_deposit_min}")
    if security_deposit_max: active_filters.append(f"Security Deposit (Max): ₹{security_deposit_max}")
    
    ws['A4'] = f"Filters Applied: {'Yes' if active_filters else 'No'}"
    if active_filters:
        ws['A5'] = "Active Filters:"
        for i, filter_text in enumerate(active_filters[:5]):  # Show first 5 filters
            ws[f'A{6+i}'] = f"  • {filter_text}"
        if len(active_filters) > 5:
            ws[f'A{6+5}'] = f"  • ... and {len(active_filters) - 5} more"
    
    # Style summary section
    summary_font = Font(bold=True, size=12)
    summary_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    
    # Calculate the last summary row
    last_summary_row = 4
    if active_filters:
        last_summary_row = 6 + min(len(active_filters), 5)  # 5 filters max shown
        if len(active_filters) > 5:
            last_summary_row += 1  # Add one more row for "and X more" message
    
    for row in range(1, last_summary_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = summary_font
        cell.fill = summary_fill
    
    # Start data from row after summary + 1 blank row
    data_start_row = last_summary_row + 2
    
    # Define headers
    headers = [
        'S.No', 'CFA Code', 'CFA Name', 'State', 'State Code', 'Branch Name', 'District Code', 
        'District Name', 'SPO Code', 'SPO Name', 'Structure Group', 'CFA Status', 'Agreement Type',
        'Inception Date', 'Agreement From Date', 'Agreement To Date', 'Godown Address', 'CFA Address',
        'Owner Name', 'Owner Contact', 'Owner Email', 'GST No', 'PAN No', 'Bank Account Name',
        'Bank Account No', 'Bank Name', 'Bank Branch', 'Bank IFSC', 'Destination Code',
        'Security Deposit', 'Security Deposit Doc', 'Closure Date', 'Remarks', 'Status'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    for row_idx, record in enumerate(records, data_start_row + 1):
        row_data = [
            row_idx - data_start_row,  # S.No
            record.cfa_code or 'N/A',
            record.cfa_name or 'N/A',
            record.state.state_name if record.state else 'N/A',
            record.state.state_code if record.state else 'N/A',
            record.branch.state_branch_name if record.branch else 'N/A',
            record.district_code or 'N/A',
            record.district.name if record.district else 'N/A',
            record.spo_code or 'N/A',
            record.spo_name or 'N/A',
            record.stru_grp or 'N/A',
            record.cfa_status or 'N/A',
            record.agreement_renewal or 'N/A',
            record.inception_date.strftime('%Y-%m-%d') if record.inception_date else 'N/A',
            record.agreement_from_date.strftime('%Y-%m-%d') if record.agreement_from_date else 'N/A',
            record.agreement_to_date.strftime('%Y-%m-%d') if record.agreement_to_date else 'N/A',
            record.godown_address or 'N/A',
            record.cfa_address or 'N/A',
            record.owner_name or 'N/A',
            record.owner_contact_no or 'N/A',
            record.cfa_mail_id or 'N/A',
            record.gst_no or 'N/A',
            record.pan_no or 'N/A',
            record.bank_account_name or 'N/A',
            record.bank_account_no or 'N/A',
            record.bank_name or 'N/A',
            record.bank_branch_name or 'N/A',
            record.bank_ifsc_code or 'N/A',
            record.destination_code or 'N/A',
            f"₹{record.security_deposit_rs:.2f}" if record.security_deposit_rs else 'N/A',
            record.security_deposit_doc_ref_dd or 'N/A',
            record.closure_date.strftime('%Y-%m-%d') if record.closure_date else 'N/A',
            record.remarks or 'N/A',
            record.status or 'N/A'
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="CFA_Agreement_Records_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save to response
    wb.save(response)
    
    # Add success message (will be shown when user returns to the list page)
    messages.success(request, f'Successfully exported {records.count()} CFA Agreement records to Excel format.')
    
    return response

@login_required
def cfa_agreement_create(request):
    if request.method == 'POST':
        form = CFAAgreementForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Handle branch and district selection
                branch_id = request.POST.get('branch')
                district_id = request.POST.get('district')
                
                if branch_id:
                    try:
                        branch = MasStateBranch.objects.get(id=branch_id)
                        form.instance.branch = branch
                        # Set district_code from branch if no district is selected
                        if not district_id:
                            form.instance.district_code = branch.state_branch_code
                    except MasStateBranch.DoesNotExist:
                        messages.warning(request, f'Selected branch (ID: {branch_id}) not found. Using form data.')
                
                if district_id:
                    try:
                        district = MasDistrict.objects.get(id=district_id)
                        form.instance.district = district
                        # Set district_code from district
                        form.instance.district_code = district.code
                    except MasDistrict.DoesNotExist:
                        messages.warning(request, f'Selected district (ID: {district_id}) not found. Using form data.')
                
                # Save the form
                cfa_agreement = form.save()
                
                # Create approval workflow for the new record
                try:
                    workflow = create_approval_workflow(
                        record_type='cfa_agreement',
                        record_id=cfa_agreement.id,
                        record_code=cfa_agreement.cfa_code or cfa_agreement.cfa_name,
                        submitted_by=request.user
                    )
                    if workflow:
                        logger.info(f"Approval workflow created for CFA Agreement record: {cfa_agreement.cfa_code or cfa_agreement.cfa_name}")
                    else:
                        logger.warning(f"Failed to create approval workflow for CFA Agreement record: {cfa_agreement.cfa_code or cfa_agreement.cfa_name}")
                except Exception as workflow_error:
                    logger.error(f"Error creating approval workflow: {str(workflow_error)}")
                
                # Send automatic email notification
                try:
                    send_cfa_agreement_creation_email(cfa_agreement, request.user.username)
                except Exception as email_error:
                    logger.error(f"Failed to send CFA creation email: {str(email_error)}")
                    # Don't fail the form submission if email fails
                
                messages.success(request, f'CFA Agreement "{cfa_agreement.cfa_name}" created successfully!')
                return redirect('cfa_agreement_list')
                
            except Exception as e:
                messages.error(request, f'Error creating CFA Agreement: {str(e)}')
                print(f"Error creating CFA Agreement: {e}")
        else:
            # Show form errors
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            
            if error_messages:
                messages.error(request, f'Please correct the following errors: {" ".join(error_messages)}')
            else:
                messages.error(request, 'Please correct the errors below.')
    else:
        form = CFAAgreementForm()
    
    return render(request, 'dashboard/cfa_agreement_form.html', {
        'form': form, 
        'title': 'Add CFA Agreement', 
        'is_edit': False
    })

@login_required
def cfa_agreement_edit(request, cfa_id):
    cfa = get_object_or_404(CFAAgreement, id=cfa_id)
    if request.method == 'POST':
        form = CFAAgreementForm(request.POST, request.FILES, instance=cfa)
        if form.is_valid():
            # Handle branch and district selection
            branch_id = request.POST.get('branch')
            district_id = request.POST.get('district')
            
            if branch_id:
                try:
                    branch = MasStateBranch.objects.get(id=branch_id)
                    form.instance.branch = branch
                    # Set district_code from branch if no district is selected
                    if not district_id:
                        form.instance.district_code = branch.state_branch_code
                except MasStateBranch.DoesNotExist:
                    pass
            
            if district_id:
                try:
                    district = MasDistrict.objects.get(id=district_id)
                    form.instance.district = district
                    # Set district_code from district
                    form.instance.district_code = district.code
                except MasDistrict.DoesNotExist:
                    pass
            
            form.save()
            
            # Send automatic email notification for CFA agreement update
            try:
                send_cfa_agreement_update_email(cfa, request.user.username)
            except Exception as email_error:
                logger.error(f"Failed to send CFA update email: {str(email_error)}")
                # Don't fail the form submission if email fails
            
            messages.success(request, 'CFA Agreement updated successfully!')
            return redirect('cfa_agreement_list')
    else:
        form = CFAAgreementForm(instance=cfa)
    return render(request, 'dashboard/cfa_agreement_form.html', {'form': form, 'title': 'Edit CFA Agreement', 'cfa': cfa, 'is_edit': True})

@login_required
def cfa_agreement_view(request, cfa_id):
    cfa = get_object_or_404(CFAAgreement, id=cfa_id)
    # Get partner details for this CFA agreement
    partners = CFAPartnerDetails.objects.filter(cfa_agreement=cfa).order_by('created_at')
    return render(request, 'dashboard/cfa_agreement_view.html', {
        'cfa': cfa,
        'partners': partners
    })

@login_required
def cfa_agreement_delete(request, cfa_id):
    cfa = get_object_or_404(CFAAgreement, id=cfa_id)
    if request.method == 'POST':
        cfa.delete()
        messages.success(request, 'CFA Agreement deleted successfully!')
        return redirect('cfa_agreement_list')
    return render(request, 'dashboard/cfa_agreement_delete.html', {'cfa': cfa})

@login_required
def cfa_agreement_pdf(request, cfa_id):
    cfa = get_object_or_404(CFAAgreement, id=cfa_id)
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="CFA_Agreement_{cfa.cfa_code}_{cfa.cfa_name}.pdf"'
    
    # Create the PDF object using BytesIO
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Create custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.darkblue
    )
    
    normal_style = styles['Normal']
    
    # Title
    elements.append(Paragraph(f"CFA Agreement Details", title_style))
    elements.append(Spacer(1, 20))
    
    # Header Information
    elements.append(Paragraph(f"<b>CFA Code:</b> {cfa.cfa_code}", normal_style))
    elements.append(Paragraph(f"<b>CFA Name:</b> {cfa.cfa_name}", normal_style))
    elements.append(Paragraph(f"<b>Generated On:</b> {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", normal_style))
    elements.append(Spacer(1, 20))
    
    # CFA Information Section
    elements.append(Paragraph("CFA Information", heading_style))
    cfa_data = [
        ['Field', 'Value'],
        ['SPO Code', cfa.spo_code or 'Not specified'],
        ['SPO Name', cfa.spo_name or 'Not specified'],
        ['Structure Group', cfa.stru_grp or 'Not specified'],
        ['CFA Status', cfa.cfa_status or 'Not specified'],
        ['Inception Date', cfa.inception_date.strftime('%B %d, %Y') if cfa.inception_date else 'Not specified'],
        ['Agreement From Date', cfa.agreement_from_date.strftime('%B %d, %Y') if cfa.agreement_from_date else 'Not specified'],
        ['Agreement To Date', cfa.agreement_to_date.strftime('%B %d, %Y') if cfa.agreement_to_date else 'Not specified'],
        ['Agreement Renewal', cfa.agreement_renewal or 'Not specified'],
        ['Godown Address', cfa.godown_address or 'Not specified'],
        ['CFA Address', cfa.cfa_address or 'Not specified'],
    ]
    
    cfa_table = Table(cfa_data, colWidths=[2*inch, 4*inch])
    cfa_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(cfa_table)
    elements.append(Spacer(1, 20))
    
    # Location Information Section
    elements.append(Paragraph("Location Information", heading_style))
    location_data = [
        ['Field', 'Value'],
        ['State', cfa.state.state_name if cfa.state else 'Not specified'],
                    ['Branch Name', cfa.branch.state_branch_name if cfa.branch else 'Not specified'],
        ['District Code', cfa.district_code or 'Not specified'],
                    ['District Name', cfa.district.name if cfa.district else 'Not specified'],
    ]
    
    location_table = Table(location_data, colWidths=[2*inch, 4*inch])
    location_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(location_table)
    elements.append(Spacer(1, 20))
    
    # Owner Information Section
    elements.append(Paragraph("Owner Information", heading_style))
    owner_data = [
        ['Field', 'Value'],
        ['Owner Name', cfa.owner_name or 'Not specified'],
        ['Owner Contact No', cfa.owner_contact_no or 'Not specified'],
        ['CFA Mail ID', cfa.cfa_mail_id or 'Not specified'],
    ]
    
    owner_table = Table(owner_data, colWidths=[2*inch, 4*inch])
    owner_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(owner_table)
    elements.append(Spacer(1, 20))
    
    # Business Details Section
    elements.append(Paragraph("Business Details", heading_style))
    business_data = [
        ['Field', 'Value'],
        ['GST No', cfa.gst_no or 'Not specified'],
        ['PAN No', cfa.pan_no or 'Not specified'],
    ]
    
    business_table = Table(business_data, colWidths=[2*inch, 4*inch])
    business_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(business_table)
    elements.append(Spacer(1, 20))
    
    # Bank Information Section
    elements.append(Paragraph("Bank Information", heading_style))
    bank_data = [
        ['Field', 'Value'],
        ['Bank Account Name', cfa.bank_account_name or 'Not specified'],
        ['Bank Account No', cfa.bank_account_no or 'Not specified'],
        ['Bank Name', cfa.bank_name or 'Not specified'],
        ['Bank Branch Name', cfa.bank_branch_name or 'Not specified'],
        ['Bank IFSC Code', cfa.bank_ifsc_code or 'Not specified'],
    ]
    
    bank_table = Table(bank_data, colWidths=[2*inch, 4*inch])
    bank_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(bank_table)
    elements.append(Spacer(1, 20))
    
    # Space Information Section
    elements.append(Paragraph("Space Information", heading_style))
    space_data = [
        ['Field', 'Value'],
        ['Destination Code', cfa.destination_code or 'Not specified'],
    ]
    
    space_table = Table(space_data, colWidths=[2*inch, 4*inch])
    space_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(space_table)
    elements.append(Spacer(1, 20))
    
    # Financial Information Section
    elements.append(Paragraph("Financial Information", heading_style))
    financial_data = [
        ['Field', 'Value'],
        ['Security Deposit Rs.', f"₹{cfa.security_deposit_rs}" if cfa.security_deposit_rs else 'Not specified'],
        ['Security Deposit Doc. / Ref. / DD', cfa.security_deposit_doc_ref_dd or 'Not specified'],
    ]
    
    financial_table = Table(financial_data, colWidths=[2*inch, 4*inch])
    financial_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(financial_table)
    elements.append(Spacer(1, 20))
    
    # Attachments Section
    elements.append(Paragraph("Attachments", heading_style))
    attachments_data = [
        ['Document Type', 'Status'],
        ['CFA Agreement Document', 'Available' if cfa.cfa_agreement else 'Not Available'],
        ['Closure Letter', 'Available' if cfa.closure_letter else 'Not Available'],
        ['Closure Acceptance Letter', 'Available' if cfa.closure_acceptance_letter else 'Not Available'],
        ['FF Letter Calculation', 'Available' if cfa.ff_letter_calc else 'Not Available'],
        ['Security Deposit Document', 'Available' if cfa.security_deposit else 'Not Available'],
        # Removed vacation_letter as it doesn't exist in CFAAgreement model
    ]
    
    attachments_table = Table(attachments_data, colWidths=[3*inch, 3*inch])
    attachments_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(attachments_table)
    elements.append(Spacer(1, 20))
    
    # Additional Information Section
    elements.append(Paragraph("Additional Information", heading_style))
    additional_data = [
        ['Field', 'Value'],
        ['Status', cfa.status or 'Not specified'],
        ['Closure Date', cfa.closure_date.strftime('%B %d, %Y') if cfa.closure_date else 'Not specified'],
        ['Created', cfa.created_at.strftime('%B %d, %Y at %I:%M %p') if cfa.created_at else 'Not specified'],
        ['Last Updated', cfa.updated_at.strftime('%B %d, %Y at %I:%M %p') if cfa.updated_at else 'Not specified'],
        ['Remarks', cfa.remarks or 'No remarks'],
    ]
    
    additional_table = Table(additional_data, colWidths=[2*inch, 4*inch])
    additional_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(additional_table)
    
    # Build the PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response

# Transporter Agreement Views
@login_required
def transporter_agreement_list(request):
    # Get filter parameters
    transporter_name = request.GET.get('transporter_name', '')
    transporter_code = request.GET.get('transporter_code', '')
    source_plant_name = request.GET.get('source_plant_name', '')
    district_name = request.GET.get('district_name', '')
    state = request.GET.get('state', '')
    owner_name = request.GET.get('owner_name', '')
    transporter_status = request.GET.get('transporter_status', '')
    multiple_plant = request.GET.get('multiple_plant', '')
    
    # Start with all records
    transporter_list = TransporterAgreement.objects.all()
    
    # Apply filters
    if transporter_name:
        transporter_list = transporter_list.filter(transporter_name__icontains=transporter_name)
    if transporter_code:
        transporter_list = transporter_list.filter(transporter_code__icontains=transporter_code)
    if source_plant_name:
        transporter_list = transporter_list.filter(source_plant_name__icontains=source_plant_name)
    if district_name:
        transporter_list = transporter_list.filter(district__name__icontains=district_name)
    if state:
        transporter_list = transporter_list.filter(state__state_name__icontains=state)
    if owner_name:
        transporter_list = transporter_list.filter(owner_managing_partner__icontains=owner_name)
    if transporter_status:
        transporter_list = transporter_list.filter(transporter_status=transporter_status)
    if multiple_plant:
        transporter_list = transporter_list.filter(operating_in_multiple_plant=multiple_plant)
    
    # Calculate statistics
    total_records = TransporterAgreement.objects.count()
    active_records = TransporterAgreement.objects.filter(status='Active').count()
    inactive_records = TransporterAgreement.objects.exclude(status='Active').count()
    expired_records = TransporterAgreement.objects.filter(status='Expired').count()
    terminated_records = TransporterAgreement.objects.filter(status='Terminated').count()
    filtered_records = transporter_list.count()
    
    context = {
        'transporter_list': transporter_list,
        'total_records': total_records,
        'active_records': active_records,
        'inactive_records': inactive_records,
        'expired_records': expired_records,
        'terminated_records': terminated_records,
        'filtered_records': filtered_records,
        'filters': {
            'transporter_name': transporter_name,
            'transporter_code': transporter_code,
            'source_plant_name': source_plant_name,
            'district_name': district_name,
            'state': state,
            'owner_name': owner_name,
            'transporter_status': transporter_status,
            'multiple_plant': multiple_plant,
        }
    }
    
    return render(request, 'dashboard/transporter_agreement_list.html', context)

@login_required
def transporter_agreement_export_excel(request):
    """Export Transporter Agreement records to Excel format"""
    # Get filtered records (same logic as transporter_agreement_list)
    transporter_name = request.GET.get('transporter_name', '')
    transporter_code = request.GET.get('transporter_code', '')
    source_plant_name = request.GET.get('source_plant_name', '')
    district_name = request.GET.get('district_name', '')
    state = request.GET.get('state', '')
    owner_name = request.GET.get('owner_name', '')
    transporter_status = request.GET.get('transporter_status', '')
    multiple_plant = request.GET.get('multiple_plant', '')
    
    # Start with all records
    records = TransporterAgreement.objects.select_related('state', 'branch').all()
    
    # Apply filters
    if transporter_name:
        records = records.filter(transporter_name__icontains=transporter_name)
    if transporter_code:
        records = records.filter(transporter_code__icontains=transporter_code)
    if source_plant_name:
        records = records.filter(source_plant_name__icontains=source_plant_name)
    if district_name:
        records = records.filter(district__name__icontains=district_name)
    if state:
        records = records.filter(state__state_name__icontains=state)
    if owner_name:
        records = records.filter(owner_managing_partner__icontains=owner_name)
    if transporter_status:
        records = records.filter(transporter_status=transporter_status)
    if multiple_plant:
        records = records.filter(operating_in_multiple_plant=multiple_plant)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transporter Agreement Records"
    
    # Add summary information
    ws['A1'] = "Transporter Agreement Records Export"
    ws['A2'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Total Records: {records.count()}"
    
    # Show active filters
    active_filters = []
    if transporter_name: active_filters.append(f"Transporter Name: {transporter_name}")
    if transporter_code: active_filters.append(f"Transporter Code: {transporter_code}")
    if source_plant_name: active_filters.append(f"Source Plant: {source_plant_name}")
    if district_name: active_filters.append(f"District: {district_name}")
    if state: active_filters.append(f"State: {state}")
    if owner_name: active_filters.append(f"Owner: {owner_name}")
    if transporter_status: active_filters.append(f"Status: {transporter_status}")
    if multiple_plant: active_filters.append(f"Multiple Plant: {multiple_plant}")
    
    ws['A4'] = f"Filters Applied: {'Yes' if active_filters else 'No'}"
    if active_filters:
        ws['A5'] = "Active Filters:"
        for i, filter_text in enumerate(active_filters[:5]):  # Show first 5 filters
            ws[f'A{6+i}'] = f"  • {filter_text}"
        if len(active_filters) > 5:
            ws[f'A{6+5}'] = f"  • ... and {len(active_filters) - 5} more"
    
    # Style summary section
    summary_font = Font(bold=True, size=12)
    summary_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    
    # Calculate the last summary row
    last_summary_row = 4
    if active_filters:
        last_summary_row = 6 + min(len(active_filters), 5)  # 5 filters max shown
        if len(active_filters) > 5:
            last_summary_row += 1  # Add one more row for "and X more" message
    
    for row in range(1, last_summary_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = summary_font
        cell.fill = summary_fill
    
    # Start data from row after summary + 1 blank row
    data_start_row = last_summary_row + 2
    
    # Define headers
    headers = [
        'S.No', 'Transporter Code', 'Transporter Name', 'State', 'State Code', 'Branch', 'Branch Code',
        'District Code', 'District Name', 'Source Plant Code', 'Source Plant Name', 'Operating in Multiple Plant',
        'List of Plant', 'Transporter Status', 'Inception Date', 'Agreement From Date', 'Agreement To Date',
        'Transporter Address', 'Owner/Managing Partner', 'Owner Contact', 'Transporter Email', 'GST No', 'PAN No',
        'Bank Account Name', 'Bank Account No', 'Bank Name', 'Bank Branch', 'Bank IFSC', 'Customer Code for Invoicing',
        'Security Deposit', 'Security Deposit Doc Ref', 'Reg Office Destination Code', 'Closure Date', 'Remark', 'Status'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    for row_idx, record in enumerate(records, data_start_row + 1):
        row_data = [
            row_idx - data_start_row,  # S.No
            record.transporter_code or 'N/A',
            record.transporter_name or 'N/A',
            record.state.state_name if record.state else 'N/A',
            record.state.state_code if record.state else 'N/A',
            record.branch.state_branch_name if record.branch else 'N/A',
            record.branch.state_branch_code if record.branch else 'N/A',
            record.district_code or 'N/A',
            record.district.name if record.district else 'N/A',
            record.source_plant_code or 'N/A',
            record.source_plant_name or 'N/A',
            record.operating_in_multiple_plant or 'N/A',
            record.list_of_plant or 'N/A',
            record.transporter_status or 'N/A',
            record.inception_date.strftime('%Y-%m-%d') if record.inception_date else 'N/A',
            record.agreement_from_date.strftime('%Y-%m-%d') if record.agreement_from_date else 'N/A',
            record.agreement_to_date.strftime('%Y-%m-%d') if record.agreement_to_date else 'N/A',
            record.transporter_address or 'N/A',
            record.owner_managing_partner or 'N/A',
            record.owner_contact_no or 'N/A',
            record.transporter_mail_id or 'N/A',
            record.gst_no or 'N/A',
            record.pan_no or 'N/A',
            record.bank_account_name or 'N/A',
            record.bank_account_no or 'N/A',
            record.bank_name or 'N/A',
            record.branch_name or 'N/A',
            record.bank_ifsc_code or 'N/A',
            record.customer_code_for_invoicing or 'N/A',
            f"₹{record.security_deposit_rs:.2f}" if record.security_deposit_rs else 'N/A',
            record.security_deposit_doc_dd or 'N/A',
            record.reg_office_destination_code or 'N/A',
            record.closure_date.strftime('%Y-%m-%d') if record.closure_date else 'N/A',
            record.remark or 'N/A',
            record.status or 'N/A'
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Transporter_Agreement_Records_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save to response
    wb.save(response)
    
    # Add success message (will be shown when user returns to the list page)
    messages.success(request, f'Successfully exported {records.count()} Transporter Agreement records to Excel format.')
    
    return response

@login_required
def transporter_agreement_create(request):
    if request.method == 'POST':
        form = TransporterAgreementForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                transporter = form.save()
                
                # Create approval workflow for the new record
                try:
                    workflow = create_approval_workflow(
                        record_type='transporter_agreement',
                        record_id=transporter.id,
                        record_code=transporter.transporter_code,
                        submitted_by=request.user
                    )
                    if workflow:
                        logger.info(f"Approval workflow created for Transporter Agreement record: {transporter.transporter_code}")
                    else:
                        logger.warning(f"Failed to create approval workflow for Transporter Agreement record: {transporter.transporter_code}")
                except Exception as workflow_error:
                    logger.error(f"Error creating approval workflow: {str(workflow_error)}")
                
                # Send automatic email notification
                try:
                    send_transporter_agreement_creation_email(transporter, request.user.username)
                except Exception as email_error:
                    logger.error(f"Failed to send Transporter Agreement creation email: {str(email_error)}")
                    # Don't fail the form submission if email fails
                
                messages.success(request, 'Transporter Agreement created successfully!')
                return render(request, 'dashboard/transporter_agreement_success.html', {
                    'transporter': transporter
                })
            except Exception as e:
                messages.error(request, f'Error creating transporter agreement: {str(e)}')
        else:
            # Get field errors for highlighting
            field_errors = {}
            for field_name, errors in form.errors.items():
                if field_name != '__all__':
                    field_errors[field_name] = errors
            messages.error(request, 'Please correct the errors below.')
    else:
        form = TransporterAgreementForm()
    
    # Get plant data for the multiselect dropdown
    try:
        from .models import MasPlant
        import json
        plants = MasPlant.objects.filter(status=1).order_by('name')
        plant_data = [{'value': plant.name, 'label': plant.name} for plant in plants]
        plant_data_json = json.dumps(plant_data)
    except Exception as e:
        # Fallback to default choices if table doesn't exist
        plant_data = [
            {'value': 'CCPL', 'label': 'CCPL'}, {'value': 'APCL', 'label': 'APCL'}, 
            {'value': 'AW01', 'label': 'AW01'}, {'value': 'DW01', 'label': 'DW01'}, 
            {'value': 'GW01', 'label': 'GW01'}, {'value': 'KW01', 'label': 'KW01'},
            {'value': 'SG01', 'label': 'SG01'}, {'value': 'PW01', 'label': 'PW01'}, 
            {'value': 'VG01', 'label': 'VG01'}, {'value': 'F022', 'label': 'F022'},
            {'value': 'NW01', 'label': 'NW01'}, {'value': 'AD01', 'label': 'AD01'},
            {'value': 'AD15', 'label': 'AD15'}, {'value': 'TN04', 'label': 'TN04'}, 
            {'value': 'TN03', 'label': 'TN03'}
        ]
        plant_data_json = json.dumps(plant_data)
    
    return render(request, 'dashboard/transporter_agreement_form.html', {
        'form': form, 
        'title': 'Add Transporter Agreement',
        'field_errors': field_errors if 'field_errors' in locals() else {},
        'plant_data': plant_data,
        'plant_data_json': plant_data_json,
    })

@login_required
def transporter_agreement_edit(request, transporter_id):
    transporter = get_object_or_404(TransporterAgreement, id=transporter_id)
    if request.method == 'POST':
        form = TransporterAgreementForm(request.POST, request.FILES, instance=transporter)
        if form.is_valid():
            form.save()
            
            # Send automatic email notification for Transporter Agreement update
            try:
                send_transporter_agreement_update_email(transporter, request.user.username)
            except Exception as email_error:
                logger.error(f"Failed to send Transporter Agreement update email: {str(email_error)}")
                # Don't fail the form submission if email fails
            
            messages.success(request, 'Transporter Agreement updated successfully!')
            return redirect('transporter_agreement_list')
    else:
        form = TransporterAgreementForm(instance=transporter)
    
    # Get plant data for the multiselect dropdown
    try:
        from .models import MasPlant
        import json
        plants = MasPlant.objects.filter(status=1).order_by('name')
        plant_data = [{'value': plant.name, 'label': plant.name} for plant in plants]
        plant_data_json = json.dumps(plant_data)
    except Exception as e:
        # Fallback to default choices if table doesn't exist
        plant_data = [
            {'value': 'CCPL', 'label': 'CCPL'}, {'value': 'APCL', 'label': 'APCL'}, 
            {'value': 'AW01', 'label': 'AW01'}, {'value': 'DW01', 'label': 'DW01'}, 
            {'value': 'GW01', 'label': 'GW01'}, {'value': 'KW01', 'label': 'KW01'},
            {'value': 'SG01', 'label': 'SG01'}, {'value': 'PW01', 'label': 'PW01'}, 
            {'value': 'VG01', 'label': 'VG01'}, {'value': 'F022', 'label': 'F022'},
            {'value': 'NW01', 'label': 'NW01'}, {'value': 'AD01', 'label': 'AD01'},
            {'value': 'AD15', 'label': 'AD15'}, {'value': 'TN04', 'label': 'TN04'}, 
            {'value': 'TN03', 'label': 'TN03'}
        ]
        plant_data_json = json.dumps(plant_data)
    
    return render(request, 'dashboard/transporter_agreement_form.html', {
        'form': form, 
        'title': 'Edit Transporter Agreement', 
        'transporter': transporter,
        'plant_data': plant_data,
        'plant_data_json': plant_data_json,
    })

@login_required
def transporter_agreement_view(request, transporter_id):
    transporter = get_object_or_404(TransporterAgreement, id=transporter_id)
    return render(request, 'dashboard/transporter_agreement_view.html', {'transporter': transporter})

@login_required
def transporter_agreement_pdf(request, transporter_id):
    transporter = get_object_or_404(TransporterAgreement, id=transporter_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Transporter_Agreement_{transporter.transporter_code}_{transporter.transporter_name}.pdf"'
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30, alignment=TA_CENTER, textColor=colors.darkblue
    )
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=12, spaceBefore=20, textColor=colors.darkblue
    )
    normal_style = styles['Normal']

    elements.append(Paragraph(f"Transporter Agreement Details", title_style))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>Transporter Code:</b> {transporter.transporter_code}", normal_style))
    elements.append(Paragraph(f"<b>Transporter Name:</b> {transporter.transporter_name}", normal_style))
    elements.append(Paragraph(f"<b>Generated On:</b> {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", normal_style))
    elements.append(Spacer(1, 20))

    # Transporter Information Section
    elements.append(Paragraph("Transporter Information", heading_style))
    transporter_data = [
        ['Field', 'Value'],
        ['Transporter Code', transporter.transporter_code or 'Not specified'],
        ['Transporter Name', transporter.transporter_name or 'Not specified'],
        ['Transporter Status', transporter.transporter_status or 'Not specified'],
        ['Inception Date', transporter.inception_date.strftime('%B %d, %Y') if transporter.inception_date else 'Not specified'],
        ['Agreement From Date', transporter.agreement_from_date.strftime('%B %d, %Y') if transporter.agreement_from_date else 'Not specified'],
        ['Agreement To Date', transporter.agreement_to_date.strftime('%B %d, %Y') if transporter.agreement_to_date else 'Not specified'],
        ['Status', transporter.status or 'Not specified'],
    ]
    transporter_table = Table(transporter_data, colWidths=[2*inch, 4*inch])
    transporter_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(transporter_table)
    elements.append(Spacer(1, 20))

    # Location Information Section
    elements.append(Paragraph("Location Information", heading_style))
    location_data = [
        ['Field', 'Value'],
        ['State', transporter.state.state_name if transporter.state else 'Not specified'],
        ['District Name', transporter.district.name if transporter.district else 'Not specified'],
        ['District Code', transporter.district_code or 'Not specified'],
    ]
    location_table = Table(location_data, colWidths=[2*inch, 4*inch])
    location_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(location_table)
    elements.append(Spacer(1, 20))

    # Plant Information Section
    elements.append(Paragraph("Plant Information", heading_style))
    plant_data = [
        ['Field', 'Value'],
        ['Source Plant Code', transporter.source_plant_code or 'Not specified'],
        ['Source Plant Name', transporter.source_plant_name or 'Not specified'],
        ['Operating in Multiple Plant', transporter.operating_in_multiple_plant or 'Not specified'],
        ['List of Plant', transporter.list_of_plant or 'Not specified'],
    ]
    plant_table = Table(plant_data, colWidths=[2*inch, 4*inch])
    plant_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(plant_table)
    elements.append(Spacer(1, 20))

    # Contact Information Section
    elements.append(Paragraph("Contact Information", heading_style))
    contact_data = [
        ['Field', 'Value'],
        ['Owner/Managing Partner', transporter.owner_managing_partner or 'Not specified'],
        ['Owner Contact No', transporter.owner_contact_no or 'Not specified'],
        ['Transporter Mail ID', transporter.transporter_mail_id or 'Not specified'],
    ]
    contact_table = Table(contact_data, colWidths=[2*inch, 4*inch])
    contact_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(contact_table)
    elements.append(Spacer(1, 20))

    # Address Information Section
    elements.append(Paragraph("Address Information", heading_style))
    address_data = [
        ['Field', 'Value'],
        ['Transporter Address', transporter.transporter_address or 'Not specified'],
    ]
    address_table = Table(address_data, colWidths=[2*inch, 4*inch])
    address_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(address_table)
    elements.append(Spacer(1, 20))

    # Business Details Section
    elements.append(Paragraph("Business Details", heading_style))
    business_data = [
        ['Field', 'Value'],
        ['GST No', transporter.gst_no or 'Not specified'],
        ['PAN No', transporter.pan_no or 'Not specified'],
    ]
    business_table = Table(business_data, colWidths=[2*inch, 4*inch])
    business_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(business_table)
    elements.append(Spacer(1, 20))

    # Bank Information Section
    elements.append(Paragraph("Bank Information", heading_style))
    bank_data = [
        ['Field', 'Value'],
        ['Bank Account Name', transporter.bank_account_name or 'Not specified'],
        ['Bank Account No', transporter.bank_account_no or 'Not specified'],
        ['Bank Name', transporter.bank_name or 'Not specified'],
        ['Bank Branch Name', transporter.branch_name or 'Not specified'],
        ['Bank IFSC Code', transporter.bank_ifsc_code or 'Not specified'],
    ]
    bank_table = Table(bank_data, colWidths=[2*inch, 4*inch])
    bank_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(bank_table)
    elements.append(Spacer(1, 20))

    # Financial Information Section
    elements.append(Paragraph("Financial Information", heading_style))
    financial_data = [
        ['Field', 'Value'],
        ['Customer Code for Invoicing', transporter.customer_code_for_invoicing or 'Not specified'],
        ['Security Deposit Rs.', f"₹{transporter.security_deposit_rs}" if transporter.security_deposit_rs else 'Not specified'],
        ['Security Deposit Doc. / DD', transporter.security_deposit_doc_dd or 'Not specified'],
        ['Reg Office Destination Code', transporter.reg_office_destination_code or 'Not specified'],
    ]
    financial_table = Table(financial_data, colWidths=[2*inch, 4*inch])
    financial_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(financial_table)
    elements.append(Spacer(1, 20))

    # Additional Information Section
    elements.append(Paragraph("Additional Information", heading_style))
    additional_data = [
        ['Field', 'Value'],
        ['Closure Date', transporter.closure_date.strftime('%B %d, %Y') if transporter.closure_date else 'Not specified'],
        ['Created', transporter.created_at.strftime('%B %d, %Y at %I:%M %p') if transporter.created_at else 'Not specified'],
        ['Last Updated', transporter.updated_at.strftime('%B %d, %Y at %I:%M %p') if transporter.updated_at else 'Not specified'],
        ['Remarks', transporter.remark or 'No remarks'],
    ]
    additional_table = Table(additional_data, colWidths=[2*inch, 4*inch])
    additional_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(additional_table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

@login_required
def transporter_agreement_delete(request, transporter_id):
    transporter = get_object_or_404(TransporterAgreement, id=transporter_id)
    if request.method == 'POST':
        transporter.delete()
        messages.success(request, 'Transporter Agreement deleted successfully!')
        return redirect('transporter_agreement_list')
    return render(request, 'dashboard/transporter_agreement_delete.html', {'transporter': transporter})



# Approval Views
@login_required
def approval_list(request):
    """List all approvals with filtering and pagination"""
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    nature_filter = request.GET.get('nature', '')
    search_query = request.GET.get('search', '')
    
    # Start with all approvals
    approvals = Approval.objects.all()
    
    # Apply filters
    if status_filter:
        approvals = approvals.filter(status=status_filter)
    
    if nature_filter:
        approvals = approvals.filter(nature_of_approval=nature_filter)
    
    if search_query:
        approvals = approvals.filter(
            Q(nfa_no__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(subject__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(approvals, 10)  # Show 10 approvals per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'records': page_obj,
        'status_filter': status_filter,
        'nature_filter': nature_filter,
        'search_query': search_query,
        'status_choices': Approval.STATUS_CHOICES,
        'nature_choices': Approval.NATURE_CHOICES,
        'total_approvals': approvals.count(),
        'pending_count': approvals.filter(status='pending').count(),
        'approved_count': approvals.filter(status='approved').count(),
        'rejected_count': approvals.filter(status='rejected').count(),
    }
    return render(request, 'dashboard/approval_list.html', context)

@login_required
def approval_export_excel(request):
    """Export Approval records to Excel format"""
    # Get filtered records (same logic as approval_list)
    status_filter = request.GET.get('status', '')
    nature_filter = request.GET.get('nature', '')
    search_query = request.GET.get('search', '')
    
    # Start with all approvals
    records = Approval.objects.all()
    
    # Apply filters
    if status_filter:
        records = records.filter(status=status_filter)
    
    if nature_filter:
        records = records.filter(nature_of_approval=nature_filter)
    
    if search_query:
        records = records.filter(
            Q(nfa_no__icontains=search_query) |
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(subject__icontains=search_query)
        )
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Approval Records"
    
    # Add summary information
    ws['A1'] = "Approval Records Export"
    ws['A2'] = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Total Records: {records.count()}"
    
    # Show active filters
    active_filters = []
    if status_filter: active_filters.append(f"Status: {status_filter}")
    if nature_filter: active_filters.append(f"Nature: {nature_filter}")
    if search_query: active_filters.append(f"Search: {search_query}")
    
    ws['A4'] = f"Filters Applied: {'Yes' if active_filters else 'No'}"
    if active_filters:
        ws['A5'] = "Active Filters:"
        for i, filter_text in enumerate(active_filters[:5]):  # Show first 5 filters
            ws[f'A{6+i}'] = f"  • {filter_text}"
        if len(active_filters) > 5:
            ws[f'A{6+5}'] = f"  • ... and {len(active_filters) - 5} more"
    
    # Style summary section
    summary_font = Font(bold=True, size=12)
    summary_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    
    # Calculate the last summary row
    last_summary_row = 4
    if active_filters:
        last_summary_row = 6 + min(len(active_filters), 5)  # 5 filters max shown
        if len(active_filters) > 5:
            last_summary_row += 1  # Add one more row for "and X more" message
    
    for row in range(1, last_summary_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.font = summary_font
        cell.fill = summary_fill
    
    # Start data from row after summary + 1 blank row
    data_start_row = last_summary_row + 2
    
    # Define headers
    headers = [
        'S.No', 'NFA No', 'NFA Date', 'Nature of Approval', 'Code', 'Name', 'Subject',
        'Valid From', 'Valid To', 'Remark', 'Status', 'Created At', 'Created By',
        'Approved By', 'Approved At', 'Updated At'
    ]
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=data_start_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    for row_idx, record in enumerate(records, data_start_row + 1):
        row_data = [
            row_idx - data_start_row,  # S.No
            record.nfa_no or 'N/A',
            record.nfa_date.strftime('%Y-%m-%d') if record.nfa_date else 'N/A',
            record.get_nature_of_approval_display() or 'N/A',
            record.code or 'N/A',
            record.name or 'N/A',
            record.subject or 'N/A',
            record.valid_from.strftime('%Y-%m-%d') if record.valid_from else 'N/A',
            record.valid_to.strftime('%Y-%m-%d') if record.valid_to else 'N/A',
            record.remark or 'N/A',
            record.get_status_display() or 'N/A',
            record.created_at.strftime('%Y-%m-%d %H:%M') if record.created_at else 'N/A',
            record.created_by.get_full_name() if record.created_by else 'N/A',
            record.approved_by.get_full_name() if record.approved_by else 'N/A',
            record.approved_at.strftime('%Y-%m-%d %H:%M') if record.approved_at else 'N/A',
            record.updated_at.strftime('%Y-%m-%d %H:%M') if record.updated_at else 'N/A'
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Approval_Records_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save to response
    wb.save(response)
    
    # Add success message (will be shown when user returns to the list page)
    messages.success(request, f'Successfully exported {records.count()} Approval records to Excel format.')
    
    return response

@login_required
def approval_create(request):
    """Create a new approval"""
    if request.method == 'POST':
        form = ApprovalForm(request.POST, request.FILES)
        if form.is_valid():
            approval = form.save(commit=False)
            approval.created_by = request.user
            approval.save()
            messages.success(request, 'Approval created successfully!')
            return redirect('approval_list')
    else:
        form = ApprovalForm()
    
    return render(request, 'dashboard/approval_form.html', {
        'form': form,
        'title': 'Create New Approval'
    })

@login_required
def approval_edit(request, approval_id):
    """Edit an existing approval"""
    approval = get_object_or_404(Approval, id=approval_id)
    
    if request.method == 'POST':
        form = ApprovalForm(request.POST, request.FILES, instance=approval)
        if form.is_valid():
            form.save()
            messages.success(request, 'Approval updated successfully!')
            return redirect('approval_list')
    else:
        form = ApprovalForm(instance=approval)
    
    return render(request, 'dashboard/approval_form.html', {
        'form': form,
        'approval': approval,
        'title': 'Edit Approval'
    })

@login_required
def approval_view(request, approval_id):
    """View approval details"""
    approval = get_object_or_404(Approval, id=approval_id)
    return render(request, 'dashboard/approval_view.html', {'approval': approval})

@login_required
def approval_delete(request, approval_id):
    """Delete an approval"""
    approval = get_object_or_404(Approval, id=approval_id)
    
    if request.method == 'POST':
        approval.delete()
        messages.success(request, 'Approval deleted successfully!')
        return redirect('approval_list')
    
    return render(request, 'dashboard/approval_delete.html', {'approval': approval})

@login_required
def approval_approve(request, approval_id):
    """Approve an approval"""
    approval = get_object_or_404(Approval, id=approval_id)
    
    if request.method == 'POST':
        if approval.can_approve():
            approval.status = 'approved'
            approval.approved_by = request.user
            approval.approved_at = timezone.now()
            approval.save()
            messages.success(request, f'Approval {approval.nfa_no} has been approved successfully!')
        else:
            messages.error(request, 'This approval cannot be approved.')
        
        return redirect('approval_list')
    
    return render(request, 'dashboard/approval_approve.html', {'approval': approval})

@login_required
def approval_reject(request, approval_id):
    """Reject an approval"""
    approval = get_object_or_404(Approval, id=approval_id)
    
    if request.method == 'POST':
        reason = request.POST.get('reason', '')
        if approval.can_reject():
            approval.status = 'rejected'
            approval.remark = f"Rejected: {reason}" if reason else "Rejected"
            approval.save()
            messages.success(request, f'Approval {approval.nfa_no} has been rejected.')
        else:
            messages.error(request, 'This approval cannot be rejected.')
        
        return redirect('approval_list')
    
    return render(request, 'dashboard/approval_reject.html', {'approval': approval})

@login_required
def approval_pdf(request, approval_id):
    approval = get_object_or_404(Approval, id=approval_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Approval_{approval.nfa_no}_{approval.name}.pdf"'
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30, alignment=TA_CENTER, textColor=colors.darkblue
    )
    heading_style = ParagraphStyle(
        'CustomHeading', parent=styles['Heading2'], fontSize=14, spaceAfter=12, spaceBefore=20, textColor=colors.darkblue
    )
    normal_style = styles['Normal']

    elements.append(Paragraph(f"Approval Details", title_style))
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"<b>NFA No:</b> {approval.nfa_no}", normal_style))
    elements.append(Paragraph(f"<b>Name:</b> {approval.name}", normal_style))
    elements.append(Paragraph(f"<b>Generated On:</b> {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", normal_style))
    elements.append(Spacer(1, 20))

    # Basic Information Section
    elements.append(Paragraph("Basic Information", heading_style))
    basic_data = [
        ['Field', 'Value'],
        ['NFA No', approval.nfa_no or 'Not specified'],
        ['NFA Date', approval.nfa_date.strftime('%B %d, %Y') if approval.nfa_date else 'Not specified'],
        ['Nature of Approval', approval.get_nature_of_approval_display() or 'Not specified'],
        ['Code', approval.code or 'Not specified'],
        ['Name', approval.name or 'Not specified'],
        ['Subject', approval.subject or 'Not specified'],
        ['Status', approval.get_status_display() or 'Not specified'],
    ]
    basic_table = Table(basic_data, colWidths=[2*inch, 4*inch])
    basic_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(basic_table)
    elements.append(Spacer(1, 20))

    # Validity Period Section
    elements.append(Paragraph("Validity Period", heading_style))
    validity_data = [
        ['Field', 'Value'],
        ['Valid From', approval.valid_from.strftime('%B %d, %Y') if approval.valid_from else 'Not specified'],
        ['Valid To', approval.valid_to.strftime('%B %d, %Y') if approval.valid_to else 'Not specified'],
    ]
    validity_table = Table(validity_data, colWidths=[2*inch, 4*inch])
    validity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(validity_table)
    elements.append(Spacer(1, 20))

    # Additional Information Section
    elements.append(Paragraph("Additional Information", heading_style))
    additional_data = [
        ['Field', 'Value'],
        ['Created', approval.created_at.strftime('%B %d, %Y at %I:%M %p') if approval.created_at else 'Not specified'],
        ['Last Updated', approval.updated_at.strftime('%B %d, %Y at %I:%M %p') if approval.updated_at else 'Not specified'],
        ['Created By', approval.created_by.get_full_name() if approval.created_by else 'Not specified'],
        ['Approved By', approval.approved_by.get_full_name() if approval.approved_by else 'Not specified'],
        ['Approved At', approval.approved_at.strftime('%B %d, %Y at %I:%M %p') if approval.approved_at else 'Not specified'],
        ['Remark', approval.remark or 'No remarks'],
    ]
    additional_table = Table(additional_data, colWidths=[2*inch, 4*inch])
    additional_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12), ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    elements.append(additional_table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

@login_required
def reports(request):
    """Reports page - for generating various reports"""
    context = {
        'title': 'Reports',
        'spo_rent_count': SPORent.objects.count(),
        'cfa_agreement_count': CFAAgreement.objects.count(),
        'transporter_agreement_count': TransporterAgreement.objects.count(),
        'total_agreements': SPORent.objects.count() + CFAAgreement.objects.count() + TransporterAgreement.objects.count(),
        'recent_reports': [
            {
                'name': 'Monthly SPO Rent Report',
                'generated_date': '2025-07-23',
                'type': 'SPO Rent',
                'status': 'completed'
            },
            {
                'name': 'CFA Agreement Summary',
                'generated_date': '2025-07-22',
                'type': 'CFA Agreement',
                'status': 'completed'
            },
            {
                'name': 'Transporter Performance Report',
                'generated_date': '2025-07-21',
                'type': 'Transporter',
                'status': 'completed'
            }
        ]
    }
    return render(request, 'dashboard/reports.html', context)

@login_required
def load_branches(request):
    """AJAX view to load districts based on selected state"""
    state_id = request.GET.get('state_id')
    if state_id:
        districts = MasStateBranch.objects.filter(state_id=state_id).values('id', 'state_branch_name', 'state_branch_code')
        return JsonResponse({'branches': list(districts)})  # Keep 'branches' key for backward compatibility
    return JsonResponse({'branches': []})

@require_http_methods(["GET"])
def load_branches_for_cfa(request):
    """Load branches based on selected state for CFA Agreement form"""
    state_id = request.GET.get('state_id')
    print(f"DEBUG: Received state_id: {state_id}")  # Debug print
    
    if state_id:
        try:
            branches = MasStateBranch.objects.filter(
                state_id=state_id
            ).values('id', 'state_branch_name', 'state_branch_code')
            print(f"DEBUG: Found {branches.count()} branches")  # Debug print
            print(f"DEBUG: Branches: {list(branches)}")  # Debug print
            return JsonResponse({'branches': list(branches)})
        except Exception as e:
            print(f"DEBUG: Error: {e}")  # Debug print
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'branches': []})

@require_http_methods(["GET"])
def get_branch_details(request):
    """Get branch details including district code for CFA Agreement form"""
    branch_id = request.GET.get('branch_id')
    if branch_id:
        try:
            branch = MasStateBranch.objects.get(id=branch_id)
            return JsonResponse({
                'district_code': branch.state_branch_code,
                'branch_name': branch.state_branch_name
            })
        except MasStateBranch.DoesNotExist:
            return JsonResponse({'error': 'Branch not found'}, status=404)
    return JsonResponse({'error': 'Branch ID required'}, status=400)

@require_http_methods(["GET"])
def load_branches_for_transporter(request):
    """Load branches for Transporter Agreement form based on state selection"""
    state_id = request.GET.get('state_id')
    if state_id:
        try:
            branches = MasStateBranch.objects.filter(state_id=state_id)
            branch_list = [{'id': branch.id, 'state_branch_name': branch.state_branch_name, 'state_branch_code': branch.state_branch_code} for branch in branches]
            return JsonResponse({'branches': branch_list})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error': 'State ID required'}, status=400)

@login_required
def mail_reminder(request):
    """Mail Reminder page - for managing email reminders"""
    # Get SPO Rent records with email addresses
    spo_rent_records = SPORent.objects.exclude(
        cfa_mail_id__isnull=True
    ).exclude(
        cfa_mail_id=''
    ).select_related('state', 'branch')
    
    # Get current date for calculations
    today = timezone.now().date()
    
    # Get statistics
    total_spo_records = SPORent.objects.count()
    records_with_email = spo_rent_records.count()
    active_records = spo_rent_records.filter(status='Active').count()
    expiring_soon = spo_rent_records.filter(
        rental_to_date__isnull=False,
        rental_to_date__lte=today + timezone.timedelta(days=30),
        rental_to_date__gte=today,
        status='Active'
    ).count()
    
    # Get expiry data for the list
    expiry_data = []
    
    # Get records expiring in next 30 days
    expiring_records = spo_rent_records.filter(
        rental_to_date__isnull=False,
        rental_to_date__lte=today + timezone.timedelta(days=30),
        rental_to_date__gte=today,
        status='Active'
    ).order_by('rental_to_date')
    
    for record in expiring_records:
        days_until_expiry = (record.rental_to_date - today).days
        expiry_data.append({
            'id': record.id,
            'spo_code': record.spo_code,
            'spo_name': record.spo_name,
            'owner_name': record.owner_name or 'N/A',
            'branch_name': record.branch.state_branch_name if record.branch else 'N/A',
            'expiry_date': record.rental_to_date,
            'days_until_expiry': days_until_expiry,
            'email': record.cfa_mail_id,
            'status': 'expiring_soon'
        })
    
    # Get expired records (past 30 days)
    expired_records = spo_rent_records.filter(
        rental_to_date__isnull=False,
        rental_to_date__lt=today,
        status='Active'
    ).order_by('-rental_to_date')[:10]  # Limit to 10 most recent
    
    for record in expired_records:
        days_expired = (today - record.rental_to_date).days
        expiry_data.append({
            'id': record.id,
            'spo_code': record.spo_code,
            'spo_name': record.spo_name,
            'owner_name': record.owner_name or 'N/A',
            'branch_name': record.branch.state_branch_name if record.branch else 'N/A',
            'expiry_date': record.rental_to_date,
            'days_until_expiry': -days_expired,
            'days_expired': days_expired,
            'email': record.cfa_mail_id,
            'status': 'expired'
        })
    
    # Get active records not expiring soon
    active_not_expiring = spo_rent_records.filter(
        rental_to_date__isnull=False,
        status='Active'
    ).exclude(
        rental_to_date__lte=today + timezone.timedelta(days=30)
    ).order_by('rental_to_date')[:5]  # Limit to 5
    
    for record in active_not_expiring:
        if record.rental_to_date:
            days_until_expiry = (record.rental_to_date - today).days
            expiry_data.append({
                'id': record.id,
                'spo_code': record.spo_code,
                'spo_name': record.spo_name,
                'owner_name': record.owner_name or 'N/A',
                'branch_name': record.branch.state_branch_name if record.branch else 'N/A',
                'expiry_date': record.rental_to_date,
                'days_until_expiry': days_until_expiry,
                'email': record.cfa_mail_id,
                'status': 'active'
            })
    
    context = {
        'title': 'Simple Mail Reminder',
        'spo_rent_records': spo_rent_records,
        'total_spo_records': total_spo_records,
        'records_with_email': records_with_email,
        'active_records': active_records,
        'expiring_soon': expiring_soon,
        'expiry_data': expiry_data,
        'debug': settings.DEBUG,
        'reminder_types': [
            {
                'name': 'SPO Rent Renewal',
                'description': 'Send reminders for SPO rent agreement renewals',
                'icon': 'fas fa-building',
                'status': 'active',
                'count': expiring_soon,
                'url': 'spo_rent_reminder'
            },
            {
                'name': 'CFA Agreement Expiry',
                'description': 'Send reminders for CFA agreement expirations',
                'icon': 'fas fa-file-contract',
                'status': 'active',
                'count': 0,
                'url': 'cfa_reminder'
            },
            {
                'name': 'Transporter Agreement Renewal',
                'description': 'Send reminders for transporter agreement renewals',
                'icon': 'fas fa-truck',
                'status': 'active',
                'count': 0,
                'url': 'transporter_reminder'
            }
        ],
        'recent_reminders': [
            {
                'type': 'SPO Rent Renewal',
                'recipient': 'john.doe@example.com',
                'sent_date': '2025-07-23',
                'status': 'sent'
            },
            {
                'type': 'CFA Agreement Expiry',
                'recipient': 'jane.smith@example.com',
                'sent_date': '2025-07-22',
                'status': 'sent'
            },
            {
                'type': 'Payment Due',
                'recipient': 'mike.wilson@example.com',
                'sent_date': '2025-07-21',
                'status': 'sent'
            }
        ]
    }
    return render(request, 'dashboard/mail_reminder.html', context)

@login_required
def spo_rent_reminder(request):
    """SPO Rent specific reminder page"""
    # Get filter parameters
    expiring_soon = request.GET.get('expiring_soon', '')
    sort_by = request.GET.get('sort_by', 'spo_code')
    order = request.GET.get('order', 'asc')
    
    # Get SPO Rent records with email addresses
    records = SPORent.objects.exclude(
        cfa_mail_id__isnull=True
    ).exclude(
        cfa_mail_id=''
    ).select_related('state', 'branch')
    
    # Apply sorting
    if sort_by == 'state':
        sort_field = 'state__state_name'
    elif sort_by == 'owner_name':
        sort_field = 'owner_name'
    elif sort_by == 'spo_name':
        sort_field = 'spo_name'
    elif sort_by == 'rental_to_date':
        sort_field = 'rental_to_date'
    elif sort_by == 'rent_pm':
        sort_field = 'rent_pm'
    else:
        sort_field = 'spo_code'
    
    # Apply order
    if order == 'desc':
        sort_field = f'-{sort_field}'
    
    records = records.order_by(sort_field)
    
    print(f"DEBUG: Found {records.count()} records with email addresses")
    for record in records:
        print(f"DEBUG: SPO Code: {record.spo_code}, Email: {record.cfa_mail_id}")
    
    # Apply filters
    if expiring_soon == 'true':
        records = records.filter(
            rental_to_date__lte=timezone.now().date() + timezone.timedelta(days=30),
            rental_to_date__gte=timezone.now().date()
        )
        print(f"DEBUG: After expiring filter: {records.count()} records")
    
    # Get statistics
    total_records = records.count()
    active_records = records.filter(status='Active').count()
    expiring_count = records.filter(
        rental_to_date__lte=timezone.now().date() + timezone.timedelta(days=30),
        rental_to_date__gte=timezone.now().date()
    ).count()
    
    print(f"DEBUG: Total records: {total_records}, Active: {active_records}, Expiring: {expiring_count}")
    
    context = {
        'title': 'SPO Rent Email Reminders',
        'records': records,
        'total_records': total_records,
        'active_records': active_records,
        'expiring_count': expiring_count,
        'today_plus_30': timezone.now().date() + timezone.timedelta(days=30),
    }
    return render(request, 'dashboard/spo_rent_reminder.html', context)

@login_required
def send_spo_rent_email(request):
    """Send email to SPO Rent users"""
    print("=" * 60)
    print("📧 SEND SPO RENT EMAIL - DEBUG")
    print("=" * 60)
    print(f"Request method: {request.method}")
    
    if request.method == 'POST':
        # Handle both JSON and form data
        if request.content_type == 'application/json':
            try:
                data = json.loads(request.body)
                spo_code = data.get('spo_code')
                email = data.get('email')
                reminder_type = data.get('reminder_type', 'renewal')
                custom_message = data.get('custom_message', '')
                
                print(f"JSON data - SPO Code: {spo_code}, Email: {email}, Type: {reminder_type}")
                
                if not spo_code:
                    return JsonResponse({
                        'success': False,
                        'message': 'SPO code is required.'
                    })
                
                # Find record by SPO code
                try:
                    record = SPORent.objects.get(spo_code=spo_code)
                except SPORent.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'message': f'SPO record with code {spo_code} not found.'
                    })
                
                if not record.cfa_mail_id:
                    return JsonResponse({
                        'success': False,
                        'message': f'No email address found for SPO {spo_code}.'
                    })
                
                                # Send email
                try:
                    # Generate email content based on type
                    if reminder_type == 'custom':
                        # Use custom subject and content from the form
                        subject = data.get('subject', f'SPO Rent Agreement Update - {record.spo_name}')
                        custom_content = data.get('content', '')
                        if custom_message:
                            custom_content += f'\n\nAdditional Message: {custom_message}'
                        
                        # Generate HTML email using custom template
                        context = {
                            'record': record,
                            'custom_message': custom_message,
                            'subject': subject,
                            'content': custom_content
                        }
                        message = render_to_string('dashboard/emails/custom_email.html', context)
                    elif reminder_type == 'renewal':
                        subject = f'SPO Rent Agreement Renewal Reminder - {record.spo_name}'
                        message = generate_renewal_email_content(record, custom_message)
                    elif reminder_type == 'expiry':
                        subject = f'SPO Rent Agreement Expiry Notice - {record.spo_name}'
                        message = generate_expiry_email_content(record, custom_message)
                    elif reminder_type == 'payment':
                        subject = f'SPO Rent Payment Reminder - {record.spo_name}'
                        message = generate_payment_email_content(record, custom_message)
                    else:
                        subject = f'SPO Rent Agreement Update - {record.spo_name}'
                        message = generate_general_email_content(record, custom_message)
                    
                    # Send HTML email using Django's EmailMessage
                    email = EmailMessage(
                        subject=subject,
                        body=message,
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        to=[record.cfa_mail_id],
                    )
                    email.content_subtype = "html"  # Set content type to HTML
                    email.send(fail_silently=False)
                    print(f"✅ Email sent successfully to {record.cfa_mail_id}: {subject}")
                    return JsonResponse({
                        'success': True,
                        'message': f'Email sent successfully to {record.cfa_mail_id}'
                    })
                    
                except Exception as email_error:
                    print(f"❌ Failed to send email to {record.cfa_mail_id}: {str(email_error)}")
                    return JsonResponse({
                        'success': False,
                        'message': f'Failed to send email: {str(email_error)}'
                    })
                    
            except json.JSONDecodeError:
                return JsonResponse({
                    'success': False,
                    'message': 'Invalid JSON data received.'
                })
        else:
            # Handle form data (existing functionality)
            record_ids = request.POST.getlist('selected_records')
            email_type = request.POST.get('email_type', 'renewal')
            custom_message = request.POST.get('custom_message', '')
            
            print(f"Form data - Selected record IDs: {record_ids}")
            print(f"Email type: {email_type}")
            print(f"Custom message: {custom_message}")
            
            if not record_ids:
                print("❌ No record IDs selected")
                messages.error(request, 'Please select at least one SPO Rent record to send email.')
                return redirect('spo_rent_reminder')
            
            # Get selected records
            records = SPORent.objects.filter(
                id__in=record_ids
            ).exclude(
                cfa_mail_id__isnull=True
            ).exclude(
                cfa_mail_id=''
            )
            
            print(f"Found {records.count()} records with valid email addresses")
            for record in records:
                print(f"  - SPO Code: {record.spo_code}, Email: {record.cfa_mail_id}")
            
            if not records.exists():
                print("❌ No records with valid email addresses found")
                messages.error(request, 'No valid email addresses found for selected records.')
                return redirect('spo_rent_reminder')
            
            # Send emails
            success_count = 0
            failed_count = 0
            
            for record in records:
                try:
                    # Generate email content based on type
                    if email_type == 'renewal':
                        subject = f'SPO Rent Agreement Renewal Reminder - {record.spo_name}'
                        message = generate_renewal_email_content(record, custom_message)
                    elif email_type == 'expiry':
                        subject = f'SPO Rent Agreement Expiry Notice - {record.spo_name}'
                        message = generate_expiry_email_content(record, custom_message)
                    elif email_type == 'payment':
                        subject = f'SPO Rent Payment Reminder - {record.spo_name}'
                        message = generate_payment_email_content(record, custom_message)
                    else:
                        subject = f'SPO Rent Agreement Update - {record.spo_name}'
                        message = generate_general_email_content(record, custom_message)
                    
                    # Send HTML email using Django's EmailMessage
                    try:
                        email = EmailMessage(
                            subject=subject,
                            body=message,
                            from_email=settings.DEFAULT_FROM_EMAIL,
                            to=[record.cfa_mail_id],
                        )
                        email.content_subtype = "html"  # Set content type to HTML
                        email.send(fail_silently=False)
                        success_count += 1
                        print(f"✅ Email sent successfully to {record.cfa_mail_id}: {subject}")
                    except Exception as email_error:
                        print(f"❌ Failed to send email to {record.cfa_mail_id}: {str(email_error)}")
                        failed_count += 1
                    
                except Exception as e:
                    print(f"Failed to send email to {record.cfa_mail_id}: {str(e)}")
                    failed_count += 1
            
            print(f"📧 Email sending complete: {success_count} successful, {failed_count} failed")
            
            if success_count > 0:
                messages.success(request, f'Successfully sent {success_count} email(s).')
            if failed_count > 0:
                messages.warning(request, f'Failed to send {failed_count} email(s).')
            
            return redirect('spo_rent_reminder')
    
    return redirect('spo_rent_reminder')

def generate_renewal_email_content(record, custom_message):
    """Generate renewal email content using HTML template"""
    context = {
        'record': record,
        'custom_message': custom_message,
        'subject': f'SPO Rent Agreement Renewal Reminder - {record.spo_name}'
    }
    return render_to_string('dashboard/emails/renewal_reminder.html', context)

def generate_expiry_email_content(record, custom_message):
    """Generate expiry email content using HTML template"""
    from datetime import date
    today = date.today()
    days_remaining = (record.rental_to_date - today).days
    
    context = {
        'record': record,
        'custom_message': custom_message,
        'days_remaining': days_remaining,
        'subject': f'SPO Rent Agreement Expiry Notice - {record.spo_name}'
    }
    return render_to_string('dashboard/emails/expiry_notice.html', context)

def generate_payment_email_content(record, custom_message):
    """Generate payment reminder email content using HTML template"""
    context = {
        'record': record,
        'custom_message': custom_message,
        'payment_due_date': 'Monthly',
        'subject': f'SPO Rent Payment Reminder - {record.spo_name}'
    }
    return render_to_string('dashboard/emails/payment_reminder.html', context)

def generate_general_email_content(record, custom_message):
    """Generate general email content using HTML template"""
    context = {
        'record': record,
        'custom_message': custom_message,
        'subject': f'SPO Rent Agreement Update - {record.spo_name}',
        'content': 'This is a general update regarding your SPO Rent Agreement.'
    }
    return render_to_string('dashboard/emails/custom_email.html', context)

@login_required
def test_email_configuration(request):
    """Test email configuration"""
    if request.method == 'POST':
        test_email = request.POST.get('test_email', '')
        
        if not test_email:
            messages.error(request, 'Please provide a test email address.')
            return redirect('mail_reminder')
        
        try:
            subject = 'Test Email from Data Management System'
            message = f"""
Hello,

This is a test email from the Data Management System to verify email configuration.

Email Details:
- From: {settings.DEFAULT_FROM_EMAIL}
- To: {test_email}
- Sent at: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}

If you receive this email, the email configuration is working correctly.

Best regards,
Chettinad Software Center
Data Management Team
            """.strip()
            
            send_mail(
                subject=subject,
                message=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[test_email],
                fail_silently=False,
            )
            
            messages.success(request, f'Test email sent successfully to {test_email}')
            print(f"✅ Test email sent successfully to {test_email}")
            
        except Exception as e:
            messages.error(request, f'Failed to send test email: {str(e)}')
            print(f"❌ Failed to send test email: {str(e)}")
        
        return redirect('mail_reminder')
    
    return redirect('mail_reminder')

@login_required
def cfa_reminder(request):
    """CFA Agreement reminder page - Coming Soon"""
    messages.info(request, 'CFA Agreement reminder functionality is coming soon!')
    return redirect('mail_reminder')

@login_required
def transporter_reminder(request):
    """Transporter Agreement reminder page - Coming Soon"""
    messages.info(request, 'Transporter Agreement reminder functionality is coming soon!')
    return redirect('mail_reminder')

@login_required
def payment_reminder(request):
    """Payment reminder page - Coming Soon"""
    messages.info(request, 'Payment reminder functionality is coming soon!')
    return redirect('mail_reminder')

@login_required
def document_reminder(request):
    """Document expiry reminder page - Coming Soon"""
    messages.info(request, 'Document expiry reminder functionality is coming soon!')
    return redirect('mail_reminder')

@login_required
def email_templates(request):
    """Email Template Manager page"""
    context = {
        'title': 'Email Template Manager',
        'debug': settings.DEBUG,
    }
    return render(request, 'dashboard/email_templates.html', context)

@login_required
def email_analytics(request):
    """Email Analytics Dashboard page"""
    context = {
        'title': 'Email Analytics',
        'debug': settings.DEBUG,
    }
    return render(request, 'dashboard/email_analytics.html', context)

@login_required
def test_email_ui(request):
    """Test Email UI page"""
    context = {
        'title': 'Test Email UI',
        'debug': settings.DEBUG,
    }
    return render(request, 'dashboard/test_email_ui.html', context)

# Partner Management Views
@login_required
def add_partner(request, spo_id):
    """Add partner to SPO record"""
    try:
        spo_record = get_object_or_404(SPORent, id=spo_id)
        
        # Check if maximum partners limit reached
        existing_partners = MasPartnerDetails.objects.filter(spo=spo_record).count()
        if existing_partners >= 5:
            return JsonResponse({
                'success': False,
                'message': 'Maximum 5 partners allowed per SPO.'
            })
        
        if request.method == 'POST':
            form = PartnerDetailsForm(request.POST, spo_id=spo_id)
            if form.is_valid():
                try:
                    partner = form.save(commit=False)
                    partner.spo = spo_record
                    partner.created_by = request.user
                    partner.save()
                    
                    # Create approval workflow for the new partner
                    try:
                        workflow = create_approval_workflow(
                            record_type='spo_partner',
                            record_id=partner.id,
                            record_code=f"{partner.name} - {spo_record.spo_code}",
                            submitted_by=request.user
                        )
                        if workflow:
                            logger.info(f"Approval workflow created for SPO Partner: {partner.name}")
                        else:
                            logger.warning(f"Failed to create approval workflow for SPO Partner: {partner.name}")
                    except Exception as workflow_error:
                        logger.error(f"Error creating approval workflow: {str(workflow_error)}")
                    
                    # Send email notification after partner is saved
                    try:
                        email_sent = send_partner_joined_email(spo_record, partner, request.user.username)
                        if email_sent:
                            logger.info(f"Partner joined email sent successfully for {partner.name}")
                        else:
                            logger.warning(f"Failed to send partner joined email for {partner.name}")
                    except Exception as e:
                        logger.error(f"Error sending partner joined email: {str(e)}")
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Partner added successfully! Email notification sent.',
                        'partner_count': MasPartnerDetails.objects.filter(spo=spo_record).count()
                    })
                except Exception as e:
                    logger.error(f"Error saving partner: {str(e)}")
                    return JsonResponse({
                        'success': False,
                        'message': f'Error saving partner: {str(e)}'
                    })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': form.errors
                })
        else:
            form = PartnerDetailsForm(spo_id=spo_id)
        
        return render(request, 'dashboard/add_partner_modal.html', {
            'form': form,
            'record': spo_record,
            'partner_count': existing_partners,
            'remaining_partners': 5 - existing_partners
        })
        
    except Exception as e:
        logger.error(f"Error adding partner: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'An error occurred while adding partner: {str(e)}'
        })

@login_required
def get_partners(request, spo_id):
    """Get partners for a specific SPO"""
    try:
        spo_record = get_object_or_404(SPORent, id=spo_id)
        partners = MasPartnerDetails.objects.filter(spo=spo_record).order_by('-created_at')
        
        partners_data = []
        for partner in partners:
            partners_data.append({
                'id': partner.id,
                'name': partner.name,
                'gender': partner.gender,
                'address': partner.address,
                'mail_id': partner.mail_id,
                'aadhar_no': partner.aadhar_no,
                'pan_no': partner.pan_no,
                'partner_join_date': partner.partner_join_date.strftime('%Y-%m-%d') if partner.partner_join_date else '',
                'partner_end_date': partner.partner_end_date.strftime('%Y-%m-%d') if partner.partner_end_date else '',
                'created_at': partner.created_at.strftime('%Y-%m-%d %H:%M') if partner.created_at else '',
            })
        
        return JsonResponse({
            'success': True,
            'partners': partners_data,
            'total_count': len(partners_data)
        })
        
    except Exception as e:
        logger.error(f"Error getting partners: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while fetching partners.'
        })



@login_required
def check_partner_limit(request, spo_id):
    """Check if SPO has reached partner limit"""
    try:
        spo_record = get_object_or_404(SPORent, id=spo_id)
        existing_partners = MasPartnerDetails.objects.filter(spo=spo_record).count()
        
        return JsonResponse({
            'success': True,
            'can_add': existing_partners < 5,
            'current_count': existing_partners,
            'max_allowed': 5
        })
        
    except Exception as e:
        logger.error(f"Error checking partner limit: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while checking partner limit.'
        })

@login_required
def partner_view(request, partner_id):
    """View partner details"""
    try:
        partner = get_object_or_404(MasPartnerDetails, id=partner_id)
        spo_record = partner.spo
        
        return render(request, 'dashboard/partner_view.html', {
            'partner': partner,
            'spo_record': spo_record
        })
        
    except Exception as e:
        logger.error(f"Error viewing partner: {str(e)}")
        messages.error(request, 'An error occurred while viewing partner details.')
        return redirect('spo_rent_list')

@login_required
def partner_edit(request, partner_id):
    """Edit partner details"""
    try:
        partner = get_object_or_404(MasPartnerDetails, id=partner_id)
        spo_record = partner.spo
        
        if request.method == 'POST':
            form = PartnerDetailsForm(request.POST, instance=partner, spo_id=spo_record.id)
            if form.is_valid():
                form.save()
                messages.success(request, 'Partner details updated successfully!')
                return redirect('spo_rent_list')
        else:
            form = PartnerDetailsForm(instance=partner, spo_id=spo_record.id)
        
        return render(request, 'dashboard/partner_edit.html', {
            'form': form,
            'partner': partner,
            'spo_record': spo_record
        })
        
    except Exception as e:
        logger.error(f"Error editing partner: {str(e)}")
        messages.error(request, 'An error occurred while editing partner details.')
        return redirect('spo_rent_list')


@login_required
def spo_partners_list(request, spo_id):
    """Show all partners for a specific SPO"""
    try:
        spo_record = get_object_or_404(SPORent, id=spo_id)
        partners = MasPartnerDetails.objects.filter(spo=spo_record).order_by('name')
        
        return render(request, 'dashboard/spo_partners_list.html', {
            'spo_record': spo_record,
            'partners': partners,
            'total_partners': partners.count()
        })
        
    except Exception as e:
        logger.error(f"Error viewing SPO partners list: {str(e)}")
        messages.error(request, 'An error occurred while viewing partner details.')
        return redirect('spo_rent_list')


# CFA Partner Management Views
@login_required
def add_cfa_partner(request, cfa_id):
    """Add partner to CFA Agreement record"""
    try:
        cfa_record = get_object_or_404(CFAAgreement, id=cfa_id)
        
        # Check if maximum partners limit reached
        existing_partners = CFAPartnerDetails.objects.filter(cfa_agreement=cfa_record).count()
        if existing_partners >= 5:
            return JsonResponse({
                'success': False,
                'message': 'Maximum 5 partners allowed per CFA Agreement.'
            })
        
        if request.method == 'POST':
            form = CFAPartnerDetailsForm(request.POST, cfa_agreement_id=cfa_id)
            if form.is_valid():
                try:
                    partner = form.save(commit=False)
                    partner.cfa_agreement = cfa_record
                    partner.created_by = request.user
                    partner.save()
                    
                    # Create approval workflow for the new partner
                    try:
                        workflow = create_approval_workflow(
                            record_type='cfa_partner',
                            record_id=partner.id,
                            record_code=f"{partner.name} - {cfa_record.cfa_code or cfa_record.cfa_name}",
                            submitted_by=request.user
                        )
                        if workflow:
                            logger.info(f"Approval workflow created for CFA Partner: {partner.name}")
                        else:
                            logger.warning(f"Failed to create approval workflow for CFA Partner: {partner.name}")
                    except Exception as workflow_error:
                        logger.error(f"Error creating approval workflow: {str(workflow_error)}")
                    
                    # Send automatic email notification for partner addition
                    try:
                        send_cfa_partner_joined_email(cfa_record, partner, request.user.username)
                    except Exception as email_error:
                        logger.error(f"Failed to send CFA partner email: {str(email_error)}")
                        # Don't fail the partner addition if email fails
                    
                    return JsonResponse({
                        'success': True,
                        'message': 'Partner added successfully!',
                        'partner_count': CFAPartnerDetails.objects.filter(cfa_agreement=cfa_record).count()
                    })
                except Exception as e:
                    logger.error(f"Error saving CFA partner: {str(e)}")
                    return JsonResponse({
                        'success': False,
                        'message': f'Error saving partner: {str(e)}'
                    })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': form.errors
                })
        else:
            form = CFAPartnerDetailsForm(cfa_agreement_id=cfa_id)
        
        return render(request, 'dashboard/add_cfa_partner_modal.html', {
            'form': form,
            'record': cfa_record,
            'partner_count': existing_partners,
            'remaining_partners': 5 - existing_partners
        })
        
    except Exception as e:
        logger.error(f"Error adding CFA partner: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'An error occurred while adding partner: {str(e)}'
        })

@login_required
def get_cfa_partners(request, cfa_id):
    """Get partners for a specific CFA Agreement"""
    try:
        cfa_record = get_object_or_404(CFAAgreement, id=cfa_id)
        partners = CFAPartnerDetails.objects.filter(cfa_agreement=cfa_record).order_by('-created_at')
        
        partners_data = []
        for partner in partners:
            partners_data.append({
                'id': partner.id,
                'name': partner.name,
                'gender': partner.gender,
                'address': partner.address,
                'mail_id': partner.mail_id,
                'aadhar_no': partner.aadhar_no,
                'pan_no': partner.pan_no,
                'partner_join_date': partner.partner_join_date.strftime('%Y-%m-%d') if partner.partner_join_date else '',
                'partner_end_date': partner.partner_end_date.strftime('%Y-%m-%d') if partner.partner_end_date else '',
                'created_at': partner.created_at.strftime('%Y-%m-%d %H:%M') if partner.created_at else '',
            })
        
        return JsonResponse({
            'success': True,
            'partners': partners_data,
            'total_count': len(partners_data)
        })
        
    except Exception as e:
        logger.error(f"Error getting CFA partners: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while fetching partners.'
        })



@login_required
def cfa_partners_list(request, cfa_id):
    """Show all partners for a specific CFA Agreement"""
    try:
        cfa_record = get_object_or_404(CFAAgreement, id=cfa_id)
        partners = CFAPartnerDetails.objects.filter(cfa_agreement=cfa_record).order_by('name')
        
        return render(request, 'dashboard/cfa_partners_list.html', {
            'cfa_record': cfa_record,
            'partners': partners,
            'total_partners': partners.count()
        })
        
    except Exception as e:
        logger.error(f"Error viewing CFA partners list: {str(e)}")
        messages.error(request, 'An error occurred while viewing partner details.')
        return redirect('cfa_agreement_list')

@login_required
def check_cfa_partner_limit(request, cfa_id):
    """Check if CFA Agreement has reached partner limit"""
    try:
        cfa_record = get_object_or_404(CFAAgreement, id=cfa_id)
        existing_partners = CFAPartnerDetails.objects.filter(cfa_agreement=cfa_record).count()
        
        return JsonResponse({
            'success': True,
            'can_add': existing_partners < 5,
            'current_count': existing_partners,
            'max_allowed': 5
        })
        
    except Exception as e:
        logger.error(f"Error checking CFA partner limit: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while checking partner limit.'
        })

@login_required
def cfa_partner_view(request, partner_id):
    """View CFA partner details"""
    try:
        partner = get_object_or_404(CFAPartnerDetails, id=partner_id)
        cfa_record = partner.cfa_agreement
        
        return render(request, 'dashboard/cfa_partner_view.html', {
            'partner': partner,
            'cfa_record': cfa_record
        })
        
    except Exception as e:
        logger.error(f"Error viewing CFA partner: {str(e)}")
        messages.error(request, 'An error occurred while viewing partner details.')
        return redirect('cfa_agreement_list')

@login_required
def cfa_partner_edit(request, partner_id):
    """Edit CFA partner details"""
    try:
        partner = get_object_or_404(CFAPartnerDetails, id=partner_id)
        cfa_record = partner.cfa_agreement
        
        if request.method == 'POST':
            form = CFAPartnerDetailsForm(request.POST, instance=partner, cfa_agreement_id=cfa_record.id)
            if form.is_valid():
                form.save()
                messages.success(request, 'Partner details updated successfully!')
                return redirect('cfa_agreement_list')
        else:
            form = CFAPartnerDetailsForm(instance=partner, cfa_agreement_id=cfa_record.id)
        
        return render(request, 'dashboard/cfa_partner_edit.html', {
            'form': form,
            'partner': partner,
            'cfa_record': cfa_record
        })
        
    except Exception as e:
        logger.error(f"Error editing CFA partner: {str(e)}")
        messages.error(request, 'An error occurred while editing partner details.')
        return redirect('cfa_agreement_list')


# Transporter Partner Management Views
@login_required
def add_transporter_partner(request, transporter_id):
    """Add a new partner to a Transporter Agreement"""
    try:
        transporter_record = get_object_or_404(TransporterAgreement, id=transporter_id)
        
        if request.method == 'POST':
            form = TransporterPartnerDetailsForm(request.POST, transporter_agreement_id=transporter_id)
            if form.is_valid():
                partner = form.save(commit=False)
                partner.transporter_agreement = transporter_record
                partner.created_by = request.user
                partner.save()
                
                # Create approval workflow for the new partner
                try:
                    workflow = create_approval_workflow(
                        record_type='transporter_partner',
                        record_id=partner.id,
                        record_code=f"{partner.name} - {transporter_record.transporter_code}",
                        submitted_by=request.user
                    )
                    if workflow:
                        logger.info(f"Approval workflow created for Transporter Partner: {partner.name}")
                    else:
                        logger.warning(f"Failed to create approval workflow for Transporter Partner: {partner.name}")
                except Exception as workflow_error:
                    logger.error(f"Error creating approval workflow: {str(workflow_error)}")
                
                # Send email notification
                try:
                    send_transporter_partner_joined_email(transporter_record, partner, request.user.username)
                except Exception as email_error:
                    logger.error(f"Failed to send transporter partner joined email: {str(email_error)}")
                
                return JsonResponse({
                    'success': True,
                    'message': 'Partner added successfully!'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': form.errors
                })
        else:
            # Check partner limit
            existing_partners = TransporterPartnerDetails.objects.filter(transporter_agreement=transporter_record).count()
            if existing_partners >= 5:
                messages.warning(request, 'Maximum 5 partners allowed per Transporter Agreement.')
                return redirect('transporter_agreement_list')
            
            form = TransporterPartnerDetailsForm(transporter_agreement_id=transporter_id)
            
            return render(request, 'dashboard/add_transporter_partner_modal.html', {
                'form': form,
                'record': transporter_record,
                'partner_count': existing_partners,
                'remaining_partners': 5 - existing_partners
            })
            
    except Exception as e:
        logger.error(f"Error adding transporter partner: {str(e)}")
        if request.method == 'POST':
            return JsonResponse({
                'success': False,
                'message': 'An error occurred while adding partner.'
            })
        else:
            messages.error(request, 'An error occurred while adding partner.')
            return redirect('transporter_agreement_list')


@login_required
def get_transporter_partners(request, transporter_id):
    """Get partners for a specific Transporter Agreement (AJAX)"""
    try:
        transporter_record = get_object_or_404(TransporterAgreement, id=transporter_id)
        partners = TransporterPartnerDetails.objects.filter(transporter_agreement=transporter_record).order_by('name')
        
        partners_data = []
        for partner in partners:
            partners_data.append({
                'id': partner.id,
                'name': partner.name,
                'gender': partner.gender,

                'email': partner.mail_id,
                'aadhar': partner.aadhar_no,
                'pan': partner.pan_no,
                'join_date': partner.partner_join_date.strftime('%Y-%m-%d') if partner.partner_join_date else '',
                'end_date': partner.partner_end_date.strftime('%Y-%m-%d') if partner.partner_end_date else '',
                'created_at': partner.created_at.strftime('%Y-%m-%d %H:%M') if partner.created_at else '',
            })
        
        return JsonResponse({
            'success': True,
            'partners': partners_data,
            'total_count': len(partners_data)
        })
        
    except Exception as e:
        logger.error(f"Error getting transporter partners: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while fetching partners.'
        })


@login_required
def transporter_partners_list(request, transporter_id):
    """Show all partners for a specific Transporter Agreement"""
    try:
        transporter_record = get_object_or_404(TransporterAgreement, id=transporter_id)
        partners = TransporterPartnerDetails.objects.filter(transporter_agreement=transporter_record).order_by('name')
        
        return render(request, 'dashboard/transporter_partners_list.html', {
            'transporter_record': transporter_record,
            'partners': partners,
            'total_partners': partners.count()
        })
        
    except Exception as e:
        logger.error(f"Error viewing transporter partners list: {str(e)}")
        messages.error(request, 'An error occurred while viewing partner details.')
        return redirect('transporter_agreement_list')


@login_required
def check_transporter_partner_limit(request, transporter_id):
    """Check if Transporter Agreement has reached partner limit"""
    try:
        transporter_record = get_object_or_404(TransporterAgreement, id=transporter_id)
        existing_partners = TransporterPartnerDetails.objects.filter(transporter_agreement=transporter_record).count()
        
        return JsonResponse({
            'success': True,
            'can_add': existing_partners < 5,
            'current_count': existing_partners,
            'max_allowed': 5
        })
        
    except Exception as e:
        logger.error(f"Error checking transporter partner limit: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'An error occurred while checking partner limit.'
        })


@login_required
def transporter_partner_view(request, partner_id):
    """View transporter partner details"""
    try:
        partner = get_object_or_404(TransporterPartnerDetails, id=partner_id)
        transporter_record = partner.transporter_agreement
        
        return render(request, 'dashboard/transporter_partner_view.html', {
            'partner': partner,
            'transporter_record': transporter_record
        })
        
    except Exception as e:
        logger.error(f"Error viewing transporter partner: {str(e)}")
        messages.error(request, 'An error occurred while viewing partner details.')
        return redirect('transporter_agreement_list')


@login_required
def transporter_partner_edit(request, partner_id):
    """Edit transporter partner details"""
    try:
        partner = get_object_or_404(TransporterPartnerDetails, id=partner_id)
        transporter_record = partner.transporter_agreement
        
        if request.method == 'POST':
            form = TransporterPartnerDetailsForm(request.POST, instance=partner, transporter_agreement_id=transporter_record.id)
            if form.is_valid():
                form.save()
                messages.success(request, 'Partner details updated successfully!')
                return redirect('transporter_agreement_list')
        else:
            form = TransporterPartnerDetailsForm(instance=partner, transporter_agreement_id=transporter_record.id)
        
        return render(request, 'dashboard/transporter_partner_edit.html', {
            'form': form,
            'partner': partner,
            'transporter_record': transporter_record
        })
        
    except Exception as e:
        logger.error(f"Error editing transporter partner: {str(e)}")
        messages.error(request, 'An error occurred while editing partner details.')
        return redirect('transporter_agreement_list')


def send_transporter_partner_joined_email(transporter_record, partner, created_by=None):
    """
    Send email notification when a new partner joins a Transporter Agreement
    """
    try:
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        subject = f'New Partner Added to Transporter Agreement - {transporter_record.transporter_code}'
        context = {
            'transporter_record': transporter_record,
            'partner': partner,
            'created_by': created_by or 'System'
        }
        html_message = render_to_string('dashboard/emails/transporter_partner_joined.html', context)
        text_message = render_to_string('dashboard/emails/transporter_partner_joined.txt', context)
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        email.content_subtype = "html"
        email.alternatives = [(text_message, 'text/plain')]
        email.send()
        logger.info(f"Transporter partner joined email sent successfully for partner ID: {partner.id}")
        return True
    except Exception as e:
        logger.error(f"Failed to send transporter partner joined email for partner ID {partner.id}: {str(e)}")
        return False

@login_required
def chatbot_questions(request):
    """Get all active chatbot questions grouped by category"""
    try:
        questions = ChatbotQuestion.objects.filter(is_active=True).order_by('order', 'question')
        
        # Group questions by category
        questions_by_category = {}
        for question in questions:
            if question.category not in questions_by_category:
                questions_by_category[question.category] = []
            questions_by_category[question.category].append({
                'id': question.id,
                'question': question.question,
                'category': question.get_category_display()
            })
        
        return JsonResponse({
            'success': True,
            'questions': questions_by_category
        })
    except Exception as e:
        logger.error(f"Error fetching chatbot questions: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch questions'
        })

@login_required
def chatbot_response(request, question_id):
    """Get response for a specific question"""
    try:
        question = get_object_or_404(ChatbotQuestion, id=question_id, is_active=True)
        
        # Create or get session
        session_id = request.session.get('chatbot_session_id')
        if not session_id:
            session_id = str(uuid.uuid4())
            request.session['chatbot_session_id'] = session_id
            session, created = ChatbotSession.objects.get_or_create(
                session_id=session_id,
                defaults={'user': request.user}
            )
        else:
            session, created = ChatbotSession.objects.get_or_create(
                session_id=session_id,
                defaults={'user': request.user}
            )
            if not created:
                session.last_activity = timezone.now()
                session.save()
        
        # Log the conversation
        ChatbotMessage.objects.create(
            session=session,
            message_type='user',
            content=f"Selected question: {question.question}",
            question=question
        )
        
        ChatbotMessage.objects.create(
            session=session,
            message_type='bot',
            content=question.response,
            question=question
        )
        
        return JsonResponse({
            'success': True,
            'response': question.response,
            'question': question.question
        })
    except Exception as e:
        logger.error(f"Error getting chatbot response: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to get response'
        })

@login_required
def chatbot_session_history(request):
    """Get conversation history for current session"""
    try:
        session_id = request.session.get('chatbot_session_id')
        if not session_id:
            return JsonResponse({
                'success': True,
                'messages': []
            })
        
        session = ChatbotSession.objects.filter(session_id=session_id).first()
        if not session:
            return JsonResponse({
                'success': True,
                'messages': []
            })
        
        messages = session.messages.all()[:50]  # Limit to last 50 messages
        
        message_list = []
        for msg in messages:
            message_list.append({
                'type': msg.message_type,
                'content': msg.content,
                'timestamp': msg.timestamp.strftime('%H:%M'),
                'question_id': msg.question.id if msg.question else None
            })
        
        return JsonResponse({
            'success': True,
            'messages': message_list
        })
    except Exception as e:
        logger.error(f"Error fetching chatbot history: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch history'
        })

@login_required
def chatbot_management(request):
    """Admin view for managing chatbot questions"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = ChatbotQuestionForm(request.POST)
        if form.is_valid():
            question = form.save()
            messages.success(request, f'Question "{question.question}" saved successfully.')
            return redirect('chatbot_management')
    else:
        form = ChatbotQuestionForm()
    
    questions = ChatbotQuestion.objects.all().order_by('order', 'question')
    
    context = {
        'form': form,
        'questions': questions,
        'categories': ChatbotQuestion._meta.get_field('category').choices
    }
    
    return render(request, 'dashboard/chatbot_management.html', context)

@login_required
def chatbot_question_edit(request, question_id):
    """Edit existing chatbot question"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    question = get_object_or_404(ChatbotQuestion, id=question_id)
    
    if request.method == 'POST':
        form = ChatbotQuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            messages.success(request, f'Question "{question.question}" updated successfully.')
            return redirect('chatbot_management')
    else:
        form = ChatbotQuestionForm(instance=question)
    
    context = {
        'form': form,
        'question': question
    }
    
    return render(request, 'dashboard/chatbot_question_edit.html', context)

@login_required
def chatbot_question_delete(request, question_id):
    """Delete chatbot question"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    question = get_object_or_404(ChatbotQuestion, id=question_id)
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, f'Question "{question.question}" deleted successfully.')
        return redirect('chatbot_management')
    
    context = {
        'question': question
    }
    
    return render(request, 'dashboard/chatbot_question_delete.html', context)

# Approval Workflow Utility Functions
def create_approval_workflow(record_type, record_id, record_code, submitted_by):
    """
    Create a new approval workflow for a record or partner
    
    Args:
        record_type (str): Type of record (e.g., 'spo_rent', 'cfa_agreement', etc.)
        record_id (int): ID of the record
        record_code (str): Code or name of the record
        submitted_by (User): User who submitted the record
    
    Returns:
        ApprovalWorkflow: Created workflow object
    """
    try:
        # Check if workflow already exists
        workflow, created = ApprovalWorkflow.objects.get_or_create(
            record_type=record_type,
            record_id=record_id,
            defaults={
                'record_code': record_code,
                'submitted_by': submitted_by,
                'status': 'waiting_confirmation'
            }
        )
        
        if created:
            # Log the creation
            ApprovalWorkflowHistory.objects.create(
                workflow=workflow,
                action='created',
                user=submitted_by,
                remarks=f'Workflow created for {record_type} record'
            )
            logger.info(f"Approval workflow created for {record_type} record {record_id}")
        else:
            # Update existing workflow if needed
            if workflow.status == 'waiting_confirmation':
                workflow.record_code = record_code
                workflow.submitted_by = submitted_by
                workflow.save()
                
                # Log the update
                ApprovalWorkflowHistory.objects.create(
                    workflow=workflow,
                    action='updated',
                    user=submitted_by,
                    remarks=f'Workflow updated for {record_type} record'
                )
                logger.info(f"Approval workflow updated for {record_type} record {record_id}")
        
        return workflow
        
    except Exception as e:
        logger.error(f"Error creating approval workflow for {record_type} record {record_id}: {str(e)}")
        return None


def get_pending_approvals(user=None, record_type=None):
    """
    Get pending approvals for approval dashboard
    
    Args:
        user (User, optional): Filter by specific user
        record_type (str, optional): Filter by record type
    
    Returns:
        QuerySet: Pending approval workflows
    """
    queryset = ApprovalWorkflow.objects.filter(
        status='waiting_confirmation',
        is_active=True
    ).select_related('submitted_by', 'approved_by')
    
    if user:
        queryset = queryset.filter(submitted_by=user)
    
    if record_type:
        queryset = queryset.filter(record_type=record_type)
    
    return queryset.order_by('-created_at')


def approve_record(workflow_id, approver, remarks=""):
    """
    Approve a record through the workflow
    
    Args:
        workflow_id (int): ID of the workflow to approve
        approver (User): User approving the record
        remarks (str): Approval remarks
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        workflow = ApprovalWorkflow.objects.get(id=workflow_id)
        
        if workflow.status != 'waiting_confirmation':
            logger.warning(f"Workflow {workflow_id} is not in waiting_confirmation status")
            return False
        
        # Approve the workflow
        workflow.approve(approver, remarks)
        
        # Log the approval
        ApprovalWorkflowHistory.objects.create(
            workflow=workflow,
            action='approved',
            user=approver,
            remarks=remarks or 'Record approved'
        )
        
        logger.info(f"Record approved: {workflow.record_type} {workflow.record_code} by {approver.username}")
        return True
        
    except ApprovalWorkflow.DoesNotExist:
        logger.error(f"Workflow {workflow_id} not found")
        return False
    except Exception as e:
        logger.error(f"Error approving workflow {workflow_id}: {str(e)}")
        return False


def reject_record(workflow_id, approver, remarks=""):
    """
    Reject a record through the workflow
    
    Args:
        workflow_id (int): ID of the workflow to reject
        approver (User): User rejecting the record
        remarks (str): Rejection remarks
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        workflow = ApprovalWorkflow.objects.get(id=workflow_id)
        
        if workflow.status != 'waiting_confirmation':
            logger.warning(f"Workflow {workflow_id} is not in waiting_confirmation status")
            return False
        
        # Reject the workflow
        workflow.reject(approver, remarks)
        
        # Log the rejection
        ApprovalWorkflowHistory.objects.create(
            workflow=workflow,
            action='rejected',
            user=approver,
            remarks=remarks or 'Record rejected'
        )
        
        logger.info(f"Record rejected: {workflow.record_type} {workflow.record_code} by {approver.username}")
        return True
        
    except ApprovalWorkflow.DoesNotExist:
        logger.error(f"Workflow {workflow_id} not found")
        return False
    except Exception as e:
        logger.error(f"Error rejecting workflow {workflow_id}: {str(e)}")
        return False

# Approval Workflow Views
@login_required
def approval_workflow_dashboard(request):
    """Dashboard for approvers to review pending records"""
    # Get pending approvals
    pending_approvals = get_pending_approvals()
    
    # Group by record type for better organization
    approvals_by_type = {}
    for approval in pending_approvals:
        record_type = approval.record_type
        if record_type not in approvals_by_type:
            approvals_by_type[record_type] = []
        approvals_by_type[record_type].append(approval)
    
    # Ensure we always have the main record types available
    main_record_types = ['spo_rent', 'cfa_agreement', 'transporter_agreement']
    for record_type in main_record_types:
        if record_type not in approvals_by_type:
            approvals_by_type[record_type] = []
    
    context = {
        'approvals_by_type': approvals_by_type,
        'total_pending': pending_approvals.count(),
        'title': 'Approval Workflow Dashboard'
    }
    
    return render(request, 'dashboard/approval_workflow_dashboard.html', context)


@login_required
def approval_workflow_detail(request, workflow_id):
    """View detailed information about a specific approval workflow"""
    try:
        workflow = ApprovalWorkflow.objects.select_related(
            'submitted_by', 'approved_by'
        ).get(id=workflow_id)
        
        # Get the actual record object
        record_obj = workflow.get_record_object()
        
        # Get workflow history
        history = workflow.history.all().order_by('-timestamp')
        
        context = {
            'workflow': workflow,
            'record_obj': record_obj,
            'history': history,
            'title': f'Approval Details - {workflow.record_code}'
        }
        
        return render(request, 'dashboard/approval_workflow_detail.html', context)
        
    except ApprovalWorkflow.DoesNotExist:
        messages.error(request, 'Approval workflow not found.')
        return redirect('approval_workflow_dashboard')


@login_required
def approval_workflow_approve(request, workflow_id):
    """Approve a record through the workflow"""
    if request.method == 'POST':
        remarks = request.POST.get('remarks', '').strip()
        
        if approve_record(workflow_id, request.user, remarks):
            messages.success(request, 'Record approved successfully!')
        else:
            messages.error(request, 'Failed to approve record. Please try again.')
        
        return redirect('approval_workflow_dashboard')
    
    # GET request - show approval form
    try:
        workflow = ApprovalWorkflow.objects.get(id=workflow_id)
        context = {
            'workflow': workflow,
            'title': f'Approve - {workflow.record_code}'
        }
        return render(request, 'dashboard/approval_workflow_approve.html', context)
        
    except ApprovalWorkflow.DoesNotExist:
        messages.error(request, 'Approval workflow not found.')
        return redirect('approval_workflow_dashboard')


@login_required
def approval_workflow_reject(request, workflow_id):
    """Reject a record through the workflow"""
    if request.method == 'POST':
        remarks = request.POST.get('remarks', '').strip()
        
        if not remarks:
            messages.error(request, 'Please provide rejection remarks.')
            return redirect('approval_workflow_reject', workflow_id)
        
        if reject_record(workflow_id, request.user, remarks):
            messages.success(request, 'Record rejected successfully!')
        else:
            messages.error(request, 'Failed to reject record. Please try again.')
        
        return redirect('approval_workflow_dashboard')
    
    # GET request - show rejection form
    try:
        workflow = ApprovalWorkflow.objects.get(id=workflow_id)
        context = {
            'workflow': workflow,
            'title': f'Reject - {workflow.record_code}'
        }
        return render(request, 'dashboard/approval_workflow_reject.html', context)
        
    except ApprovalWorkflow.DoesNotExist:
        messages.error(request, 'Approval workflow not found.')
        return redirect('approval_workflow_dashboard')


@login_required
def my_submissions(request):
    """View user's own submitted records"""
    user_submissions = ApprovalWorkflow.objects.filter(
        submitted_by=request.user,
        is_active=True
    ).select_related('approved_by').order_by('-created_at')
    
    context = {
        'submissions': user_submissions,
        'title': 'My Submissions'
    }
    
    return render(request, 'dashboard/my_submissions.html', context)


# AJAX endpoints for approval workflow
@login_required
def get_workflow_status(request, record_type, record_id):
    """Get the current status of a workflow for a specific record"""
    try:
        workflow = ApprovalWorkflow.objects.get(
            record_type=record_type,
            record_id=record_id
        )
        
        return JsonResponse({
            'status': workflow.status,
            'status_display': workflow.get_status_display(),
            'submitted_at': workflow.submitted_at.isoformat() if workflow.submitted_at else None,
            'approved_at': workflow.approved_at.isoformat() if workflow.approved_at else None,
            'approver_remarks': workflow.approver_remarks,
            'approved_by': workflow.approved_by.username if workflow.approved_by else None
        })
        
    except ApprovalWorkflow.DoesNotExist:
        return JsonResponse({
            'status': 'not_found',
            'message': 'No workflow found for this record'
        })

@csrf_exempt
def spo_records_ajax(request):
    """AJAX endpoint to get SPO records for the approval workflow dashboard"""
    try:
        logger.info("SPO records AJAX endpoint called")
        
        # Get all SPO records with related data
        spo_records = SPORent.objects.select_related(
            'state', 'branch'
        ).all().order_by('-created_at')
        
        logger.info(f"Found {spo_records.count()} SPO records")
        
        records_data = []
        for record in spo_records:
            try:
                record_data = {
                    'id': record.id,
                    'spo_code': record.spo_code,
                    'spo_name': record.spo_name,
                    'state_name': record.state.state_name if record.state else None,
                    'branch_name': record.branch.state_branch_name if record.branch else None,
                    'owner_name': record.owner_name,
                    'status': record.status,
                    'created_at': record.created_at.isoformat() if record.created_at else None,
                }
                records_data.append(record_data)
                logger.info(f"Processed SPO record: {record.spo_code}")
            except Exception as record_error:
                logger.error(f"Error processing SPO record {record.id}: {str(record_error)}")
                # Add basic record data even if there's an error with related fields
                records_data.append({
                    'id': record.id,
                    'spo_code': getattr(record, 'spo_code', 'N/A'),
                    'spo_name': getattr(record, 'spo_name', 'N/A'),
                    'state_name': 'Error',
                    'branch_name': 'Error',
                    'owner_name': getattr(record, 'owner_name', 'N/A'),
                    'status': getattr(record, 'status', 'N/A'),
                    'created_at': None,
                })
        
        logger.info(f"Successfully processed {len(records_data)} SPO records")
        
        return JsonResponse({
            'success': True,
            'records': records_data,
            'total_count': len(records_data)
        })
        
    except Exception as e:
        logger.error(f"Error loading SPO records: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'Error loading SPO records',
            'error': str(e)
        }, status=500)

@csrf_exempt
def approve_spo_record_ajax(request, record_id):
    """AJAX endpoint to approve an SPO record"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        # Get the SPO record
        spo_record = get_object_or_404(SPORent, id=record_id)
        
        # Check if there's already an approval workflow
        workflow, created = ApprovalWorkflow.objects.get_or_create(
            record_type='spo_rent',
            record_id=record_id,
            defaults={
                'record_code': spo_record.spo_code,
                'status': 'waiting_confirmation',
                'submitted_by': request.user,
            }
        )
        
        if not created and workflow.status == 'confirmed':
            return JsonResponse({
                'success': False,
                'message': 'Record is already approved'
            })
        
        # Approve the record
        workflow.status = 'confirmed'
        workflow.approved_by = request.user
        workflow.approved_at = timezone.now()
        workflow.approver_remarks = 'Approved via dashboard'
        workflow.save()
        
        # Create workflow history
        ApprovalWorkflowHistory.objects.create(
            workflow=workflow,
            action='approved',
            user=request.user,
            remarks='Approved via dashboard'
        )
        
        # Update SPO record status if needed
        if spo_record.status != 'Active':
            spo_record.status = 'Active'
            spo_record.save()
        
        # Send approval email notification
        try:
            send_spo_approval_email(spo_record, request.user)
        except Exception as email_error:
            logger.warning(f"Failed to send SPO approval email: {str(email_error)}")
        
        logger.info(f"SPO record {spo_record.spo_code} approved by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': 'SPO record approved successfully',
            'workflow_id': workflow.id,
            'status': workflow.status
        })
        
    except SPORent.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'SPO record not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error approving SPO record {record_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'Error approving record',
            'error': str(e)
        }, status=500)

@csrf_exempt
def transporter_records_ajax(request):
    """AJAX endpoint to get Transporter records for the approval workflow dashboard"""
    try:
        logger.info("Transporter records AJAX endpoint called")
        
        # Simple approach - get basic records first
        transporter_records = TransporterAgreement.objects.all()
        logger.info(f"Found {transporter_records.count()} Transporter records")
        
        records_data = []
        for record in transporter_records:
            try:
                # Get state name safely
                state_name = None
                if hasattr(record, 'state') and record.state:
                    state_name = record.state.state_name
                
                # Get branch name safely
                branch_name = None
                if hasattr(record, 'branch') and record.branch:
                    branch_name = record.branch.state_branch_name
                
                record_data = {
                    'id': record.id,
                    'transporter_code': getattr(record, 'source_plant_code', 'N/A'),  # Use source_plant_code as transporter code
                    'transporter_name': getattr(record, 'source_plant_name', 'N/A'),  # Use source_plant_name as transporter name
                    'state_name': state_name or 'N/A',
                    'branch_name': branch_name or 'N/A',
                    'source_plant_name': record.source_plant_name or 'N/A',
                    'status': record.transporter_status or 'N/A',
                    'created_at': record.created_at.isoformat() if record.created_at else None,
                }
                records_data.append(record_data)
                logger.info(f"Processed Transporter record: {record.source_plant_name}")
            except Exception as record_error:
                logger.error(f"Error processing Transporter record {record.id}: {str(record_error)}")
                # Add basic record data even if there's an error with related fields
                records_data.append({
                    'id': record.id,
                    'transporter_code': 'Error',
                    'transporter_name': 'Error',
                    'state_name': 'Error',
                    'branch_name': 'Error',
                    'source_plant_name': 'Error',
                    'status': 'Error',
                    'created_at': None,
                })
        
        logger.info(f"Successfully processed {len(records_data)} Transporter records")
        
        return JsonResponse({
            'success': True,
            'records': records_data,
            'total_count': len(records_data)
        })
        
    except Exception as e:
        logger.error(f"Error loading Transporter records: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': 'Error loading Transporter records',
            'error': str(e)
        }, status=500)

@csrf_exempt
def approve_transporter_record_ajax(request, record_id):
    """AJAX endpoint to approve a Transporter record"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        # Get the Transporter record
        transporter_record = get_object_or_404(TransporterAgreement, id=record_id)
        
        # Check if there's already an approval workflow
        workflow, created = ApprovalWorkflow.objects.get_or_create(
            record_type='transporter_agreement',
            record_id=record_id,
            defaults={
                'record_code': transporter_record.source_plant_name,
                'status': 'waiting_confirmation',
                'submitted_by': request.user,
            }
        )
        
        if not created and workflow.status == 'confirmed':
            return JsonResponse({
                'success': False,
                'message': 'Record is already approved'
            })
        
        # Approve the record
        workflow.status = 'confirmed'
        workflow.approved_by = request.user
        workflow.approved_at = timezone.now()
        workflow.approver_remarks = 'Approved via dashboard'
        workflow.save()
        
        # Create workflow history
        ApprovalWorkflowHistory.objects.create(
            workflow=workflow,
            action='approved',
            user=request.user,
            remarks='Approved via dashboard'
        )
        
        # Update Transporter record status if needed
        if transporter_record.transporter_status != 'Active':
            transporter_record.transporter_status = 'Active'
            transporter_record.save()
        
        # Send approval email notification
        try:
            send_transporter_approval_email(transporter_record, request.user)
        except Exception as email_error:
            logger.warning(f"Failed to send Transporter approval email: {str(email_error)}")
        
        logger.info(f"Transporter record {transporter_record.source_plant_name} approved by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': 'Transporter record approved successfully',
            'workflow_id': workflow.id,
            'status': workflow.status
        })
        
    except TransporterAgreement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Transporter record not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error approving Transporter record {record_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'Error approving record',
            'error': str(e)
        }, status=500)

@csrf_exempt
def cfa_records_ajax(request):
    """AJAX endpoint to get CFA records for the approval workflow dashboard"""
    try:
        logger.info("CFA records AJAX endpoint called")
        
        # Simple approach - get basic records first
        cfa_records = CFAAgreement.objects.all()
        logger.info(f"Found {cfa_records.count()} CFA records")
        
        records_data = []
        for record in cfa_records:
            try:
                # Get state name safely
                state_name = None
                if hasattr(record, 'state') and record.state:
                    state_name = record.state.state_name
                
                record_data = {
                    'id': record.id,
                    'cfa_code': record.cfa_code or 'N/A',
                    'cfa_name': record.cfa_name or 'N/A',
                    'state_name': state_name or 'N/A',
                    'branch_name': record.branch.state_branch_name if record.branch else 'N/A',
                    'owner_name': record.owner_name or 'N/A',
                    'status': record.status or 'N/A',
                    'created_at': record.created_at.isoformat() if record.created_at else None,
                }
                records_data.append(record_data)
                logger.info(f"Processed CFA record: {record.cfa_code}")
            except Exception as record_error:
                logger.error(f"Error processing CFA record {record.id}: {str(record_error)}")
                # Add basic record data even if there's an error with related fields
                records_data.append({
                    'id': record.id,
                    'cfa_code': 'Error',
                    'cfa_name': 'Error',
                    'state_name': 'Error',
                    'branch_name': 'Error',
                    'owner_name': 'Error',
                    'status': 'Error',
                    'created_at': None,
                })
        
        logger.info(f"Successfully processed {len(records_data)} CFA records")
        
        return JsonResponse({
            'success': True,
            'records': records_data,
            'total_count': len(records_data)
        })
        
    except Exception as e:
        logger.error(f"Error loading CFA records: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': 'Error loading CFA records',
            'error': str(e)
        }, status=500)

@csrf_exempt
def approve_cfa_record_ajax(request, record_id):
    """AJAX endpoint to approve a CFA record"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Invalid request method'}, status=405)
    
    try:
        # Get the CFA record
        cfa_record = get_object_or_404(CFAAgreement, id=record_id)
        
        # Check if there's already an approval workflow
        workflow, created = ApprovalWorkflow.objects.get_or_create(
            record_type='cfa_agreement',
            record_id=record_id,
            defaults={
                'record_code': cfa_record.cfa_code,
                'status': 'waiting_confirmation',
                'submitted_by': request.user,
            }
        )
        
        if not created and workflow.status == 'confirmed':
            return JsonResponse({
                'success': False,
                'message': 'Record is already approved'
            })
        
        # Approve the record
        workflow.status = 'confirmed'
        workflow.approved_by = request.user
        workflow.approved_at = timezone.now()
        workflow.approver_remarks = 'Approved via dashboard'
        workflow.save()
        
        # Create workflow history
        ApprovalWorkflowHistory.objects.create(
            workflow=workflow,
            action='approved',
            user=request.user,
            remarks='Approved via dashboard'
        )
        
        # Update CFA record status if needed
        if cfa_record.status != 'Active':
            cfa_record.status = 'Active'
            cfa_record.save()
        
        # Send approval email notification
        try:
            send_cfa_approval_email(cfa_record, request.user)
        except Exception as email_error:
            logger.warning(f"Failed to send CFA approval email: {str(email_error)}")
        
        logger.info(f"CFA record {cfa_record.cfa_code} approved by {request.user.username}")
        
        return JsonResponse({
            'success': True,
            'message': 'CFA record approved successfully',
            'workflow_id': workflow.id,
            'status': workflow.status
        })
        
    except CFAAgreement.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'CFA record not found'
        }, status=404)
    except Exception as e:
        logger.error(f"Error approving CFA record {record_id}: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': 'Error approving record',
            'error': str(e)
        }, status=500)

# Email functions for Approval Workflow Dashboard
def send_spo_approval_email(spo_record, approver):
    """
    Send email notification when an SPO record is approved
    """
    try:
        # Email recipients - you can customize these
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'SPO Record Approved - {spo_record.spo_code} ({spo_record.spo_name})'
        
        # Render email templates
        context = {
            'spo_record': spo_record,
            'approver_name': approver.get_full_name() or approver.username,
            'approver_email': approver.email,
            'approval_date': timezone.now().strftime('%B %d, %Y at %I:%M %p')
        }
        
        html_message = render_to_string('dashboard/emails/spo_approval_email.html', context)
        
        # Create email message
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Send email
        email.send()
        
        logger.info(f"SPO approval email sent successfully for record ID: {spo_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send SPO approval email for record ID {spo_record.id}: {str(e)}")
        return False

def send_cfa_approval_email(cfa_record, approver):
    """
    Send email notification when a CFA record is approved
    """
    try:
        # Email recipients - you can customize these
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'CFA Record Approved - {cfa_record.cfa_code} ({cfa_record.cfa_name})'
        
        # Render email templates
        context = {
            'cfa_record': cfa_record,
            'approver_name': approver.get_full_name() or approver.username,
            'approver_email': approver.email,
            'approval_date': timezone.now().strftime('%B %d, %Y at %I:%M %p')
        }
        
        html_message = render_to_string('dashboard/emails/cfa_approval_email.html', context)
        
        # Create email message
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Send email
        email.send()
        
        logger.info(f"CFA approval email sent successfully for record ID: {cfa_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send CFA approval email for record ID {cfa_record.id}: {str(e)}")
        return False

def send_transporter_approval_email(transporter_record, approver):
    """
    Send email notification when a Transporter record is approved
    """
    try:
        # Email recipients - you can customize these
        to_email = 'shyad.sdc@chettinad.com'
        cc_email = 'manimaran@clplho.com'
        
        # Email subject
        subject = f'Transporter Record Approved - {transporter_record.transporter_code} ({transporter_record.transporter_name})'
        
        # Render email templates
        context = {
            'transporter_record': transporter_record,
            'approver_name': approver.get_full_name() or approver.username,
            'approver_email': approver.email,
            'approval_date': timezone.now().strftime('%B %d, %Y at %I:%M %p')
        }
        
        html_message = render_to_string('dashboard/emails/transporter_approval_email.html', context)
        
        # Create email message
        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
            cc=[cc_email],
        )
        
        # Set content type to HTML
        email.content_subtype = "html"
        
        # Send email
        email.send()
        
        logger.info(f"Transporter approval email sent successfully for record ID: {transporter_record.id}")
        return True
        
    except Exception as e:
        logger.error(f"Failed to send Transporter approval email for record ID {transporter_record.id}: {str(e)}")
        return False

# Simple User Menu Control System
@login_required
def menu_access_dashboard(request):
    """Simple dashboard for controlling user menu access"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    # Get all users
    users = User.objects.all().order_by('username')
    
    # Get user menu access records
    user_menu_access = {}
    for user in users:
        # Get or create menu access for user
        menu_access, created = UserMenuAccess.objects.get_or_create(
            user=user,
            defaults={
                'spo_menu_enabled': True,
                'cfa_menu_enabled': True,
                'transport_menu_enabled': True,
                'approval_menu_enabled': True,
                'approval_workflow_enabled': True,
                'report_enabled': True,
                'mail_reminder_enabled': True,
                'user_management_enabled': True,
                'user_menu_access_control_enabled': True,
            }
        )
        
        user_menu_access[user.id] = {
            'spo_menu': menu_access.spo_menu_enabled,
            'cfa_menu': menu_access.cfa_menu_enabled,
            'transport_menu': menu_access.transport_menu_enabled,
            'approval_menu': menu_access.approval_menu_enabled,
            'approval_workflow': menu_access.approval_workflow_enabled,
            'report': menu_access.report_enabled,
            'mail_reminder': menu_access.mail_reminder_enabled,
            'user_management': menu_access.user_management_enabled,
            'user_menu_access_control': menu_access.user_menu_access_control_enabled,
        }
    
    context = {
        'users': users,
        'user_menu_access': user_menu_access,
    }
    
    return render(request, 'dashboard/simple_menu_access.html', context)

@login_required
def update_user_menu_access(request):
    """Update user menu access permissions"""
    if not request.user.is_staff:
        return JsonResponse({'success': False, 'message': 'Access denied'})
    
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        menu_type = request.POST.get('menu_type')
        enabled = request.POST.get('enabled') == 'true'
        
        try:
            user = User.objects.get(id=user_id)
            # Get or create menu access for user
            menu_access, created = UserMenuAccess.objects.get_or_create(
                user=user,
                defaults={
                    'spo_menu_enabled': True,
                    'cfa_menu_enabled': True,
                    'transport_menu_enabled': True,
                    'approval_menu_enabled': True,
                    'approval_workflow_enabled': True,
                    'report_enabled': True,
                    'mail_reminder_enabled': True,
                    'user_management_enabled': True,
                    'user_menu_access_control_enabled': True,
                }
            )
            
            # Update the specific menu access
            if hasattr(menu_access, menu_type):
                setattr(menu_access, menu_type, enabled)
                menu_access.save()
                return JsonResponse({'success': True, 'message': f'{menu_type} updated successfully'})
            else:
                return JsonResponse({'success': False, 'message': 'Invalid menu type'})
        except User.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'User not found'})
    
    return JsonResponse({'success': False, 'message': 'Invalid request method'})

# Old complex menu access control views removed - replaced with simple user menu control system
    
    return render(request, 'dashboard/role_form.html', {'form': form, 'title': 'Create Role'})

@login_required
def role_edit(request, role_id):
    """Edit an existing role"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    try:
        role = Role.objects.get(id=role_id)
        if request.method == 'POST':
            form = RoleForm(request.POST, instance=role)
            if form.is_valid():
                form.save()
                messages.success(request, 'Role updated successfully.')
                return redirect('role_list')
        else:
            form = RoleForm(instance=role)
        
        return render(request, 'dashboard/role_form.html', {'form': form, 'title': 'Edit Role', 'role': role})
    except Role.DoesNotExist:
        messages.error(request, 'Role not found.')
        return redirect('role_list')

@login_required
def role_delete(request, role_id):
    """Delete a role"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    try:
        role = Role.objects.get(id=role_id)
        if request.method == 'POST':
            role.delete()
            messages.success(request, 'Role deleted successfully.')
            return redirect('role_list')
        
        return render(request, 'dashboard/role_delete.html', {'role': role})
    except Role.DoesNotExist:
        messages.error(request, 'Role not found.')
        return redirect('role_list')

@login_required
def user_menu_access_list(request):
    """List all user menu access permissions"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    user_access = UserMenuAccess.objects.select_related('user', 'menu_item', 'role').all()
    return render(request, 'dashboard/user_menu_access_list.html', {'user_access': user_access})

@login_required
def user_menu_access_create(request):
    """Create new user menu access"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = UserMenuAccessForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'User menu access created successfully.')
            return redirect('user_menu_access_list')
    else:
        form = UserMenuAccessForm()
    
    return render(request, 'dashboard/user_menu_access_form.html', {'form': form, 'title': 'Create User Menu Access'})

@login_required
def user_menu_access_edit(request, access_id):
    """Edit existing user menu access"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    try:
        user_access = UserMenuAccess.objects.get(id=access_id)
        if request.method == 'POST':
            form = UserMenuAccessForm(request.POST, instance=user_access)
            if form.is_valid():
                form.save()
                messages.success(request, 'User menu access updated successfully.')
                return redirect('user_menu_access_list')
        else:
            form = UserMenuAccessForm(instance=user_access)
        
        return render(request, 'dashboard/user_menu_access_form.html', {'form': form, 'title': 'Edit User Menu Access', 'user_access': user_access})
    except UserMenuAccess.DoesNotExist:
        messages.error(request, 'User menu access not found.')
        return redirect('user_menu_access_list')

@login_required
def user_menu_access_delete(request, access_id):
    """Delete user menu access"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    try:
        user_access = UserMenuAccess.objects.get(id=access_id)
        if request.method == 'POST':
            user_access.delete()
            messages.success(request, 'User menu access deleted successfully.')
            return redirect('user_menu_access_list')
        
        return render(request, 'dashboard/user_menu_access_delete.html', {'user_access': user_access})
    except UserMenuAccess.DoesNotExist:
        messages.error(request, 'User menu access not found.')
        return redirect('user_menu_access_list')

@login_required
def bulk_menu_access(request):
    """Bulk assign menu access to multiple users"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = BulkMenuAccessForm(request.POST)
        if form.is_valid():
            users = form.cleaned_data['users']
            menu_items = form.cleaned_data['menu_items']
            role = form.cleaned_data['role']
            permissions = form.cleaned_data['permissions']
            
            # Create or update user menu access for each user and menu item combination
            created_count = 0
            updated_count = 0
            
            for user in users:
                for menu_item in menu_items:
                    user_access, created = UserMenuAccess.objects.get_or_create(
                        user=user,
                        menu_item=menu_item,
                        defaults={
                            'role': role,
                            'can_view': 'can_view' in permissions,
                            'can_create': 'can_create' in permissions,
                            'can_edit': 'can_edit' in permissions,
                            'can_delete': 'can_delete' in permissions,
                            'can_approve': 'can_approve' in permissions,
                        }
                    )
                    
                    if not created:
                        # Update existing access
                        user_access.role = role
                        user_access.can_view = 'can_view' in permissions
                        user_access.can_create = 'can_create' in permissions
                        user_access.can_edit = 'can_edit' in permissions
                        user_access.can_delete = 'can_delete' in permissions
                        user_access.can_approve = 'can_approve' in permissions
                        user_access.save()
                        updated_count += 1
                    else:
                        created_count += 1
            
            messages.success(request, f'Successfully assigned menu access: {created_count} new, {updated_count} updated.')
            return redirect('user_menu_access_list')
    else:
        form = BulkMenuAccessForm()
    
    return render(request, 'dashboard/bulk_menu_access.html', {'form': form})

@login_required
def user_role_list(request):
    """List all user role assignments"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    user_roles = UserRole.objects.select_related('user', 'role', 'assigned_by').all()
    return render(request, 'dashboard/user_role_list.html', {'user_roles': user_roles})

@login_required
def user_role_create(request):
    """Create new user role assignment"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = UserRoleForm(request.POST)
        if form.is_valid():
            user_role = form.save(commit=False)
            user_role.assigned_by = request.user
            user_role.save()
            messages.success(request, 'User role assigned successfully.')
            return redirect('user_role_list')
    else:
        form = UserRoleForm()
    
    return render(request, 'dashboard/user_role_form.html', {'form': form, 'title': 'Assign User Role'})

@login_required
def user_role_edit(request, user_role_id):
    """Edit existing user role assignment"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    try:
        user_role = UserRole.objects.get(id=user_role_id)
        if request.method == 'POST':
            form = UserRoleForm(request.POST, instance=user_role)
            if form.is_valid():
                form.save()
                messages.success(request, 'User role updated successfully.')
                return redirect('user_role_list')
        else:
            form = UserRoleForm(instance=user_role)
        
        return render(request, 'dashboard/user_role_form.html', {'form': form, 'title': 'Edit User Role', 'user_role': user_role})
    except UserRole.DoesNotExist:
        messages.error(request, 'User role not found.')
        return redirect('user_role_list')

@login_required
def user_role_delete(request, user_role_id):
    """Delete user role assignment"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Staff privileges required.')
        return redirect('dashboard')
    
    try:
        user_role = UserRole.objects.get(id=user_role_id)
        if request.method == 'POST':
            user_role.delete()
            messages.success(request, 'User role assignment deleted successfully.')
            return redirect('user_role_list')
        
        return render(request, 'dashboard/user_role_delete.html', {'user_role': user_role})
    except UserRole.DoesNotExist:
        messages.error(request, 'User role not found.')
        return redirect('user_role_list')

# MasDistrict Management Views
@login_required
def mas_district_list(request):
    """List all districts"""
    districts = MasDistrict.objects.select_related('mas_state', 'mas_branch').all()
    return render(request, 'dashboard/mas_district_list.html', {'districts': districts})

@login_required
def mas_district_create(request):
    """Create a new district"""
    if request.method == 'POST':
        form = MasDistrictForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'District created successfully.')
            return redirect('mas_district_list')
    else:
        form = MasDistrictForm()
    
    return render(request, 'dashboard/mas_district_form.html', {'form': form, 'title': 'Create District'})

@login_required
def mas_district_edit(request, district_id):
    """Edit an existing district"""
    try:
        district = MasDistrict.objects.get(id=district_id)
        if request.method == 'POST':
            form = MasDistrictForm(request.POST, instance=district)
            if form.is_valid():
                form.save()
                messages.success(request, 'District updated successfully.')
                return redirect('mas_district_list')
        else:
            form = MasDistrictForm(instance=district)
        
        return render(request, 'dashboard/mas_district_form.html', {'form': form, 'title': 'Edit District', 'district': district})
    except MasDistrict.DoesNotExist:
        messages.error(request, 'District not found.')
        return redirect('mas_district_list')

@login_required
def mas_district_delete(request, district_id):
    """Delete a district"""
    try:
        district = MasDistrict.objects.get(id=district_id)
        if request.method == 'POST':
            district.delete()
            messages.success(request, 'District deleted successfully.')
            return redirect('mas_district_list')
        
        return render(request, 'dashboard/mas_district_delete.html', {'district': district})
    except MasDistrict.DoesNotExist:
        messages.error(request, 'District not found.')
        return redirect('mas_district_list')

@login_required
def get_branches_by_state(request):
    """AJAX endpoint to get branches for a selected state"""
    if request.method == 'GET':
        state_id = request.GET.get('state_id')
        print(f"get_branches_by_state called with state_id: {state_id}")
        if state_id:
            try:
                branches = MasStateBranch.objects.filter(state_id=state_id).values('id', 'state_branch_name', 'state_branch_code')
                print(f"Found {branches.count()} branches for state_id={state_id}")
                return JsonResponse({'branches': list(branches)})
            except Exception as e:
                print(f"Error in get_branches_by_state: {e}")
                return JsonResponse({'error': str(e)}, status=400)
        return JsonResponse({'branches': []})
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def get_districts_by_branch(request):
    """AJAX endpoint to get districts for a selected state and branch"""
    if request.method == 'GET':
        state_id = request.GET.get('state_id')
        branch_id = request.GET.get('branch_id')
        if state_id and branch_id:
            try:
                districts = MasDistrict.objects.filter(
                    mas_state_id=state_id,
                    mas_branch_id=branch_id,
                    status=1  # Only active districts
                ).values('id', 'name', 'code')
                print(f"Found {districts.count()} districts for state_id={state_id}, branch_id={branch_id}")
                return JsonResponse({'districts': list(districts)})
            except Exception as e:
                print(f"Error in get_districts_by_branch: {e}")
                return JsonResponse({'error': str(e)}, status=400)
        return JsonResponse({'error': 'Missing state_id or branch_id'}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def get_districts_by_state(request):
    """AJAX endpoint to get districts for a selected state (without branch)"""
    if request.method == 'GET':
        state_id = request.GET.get('state_id')
        if state_id:
            try:
                districts = MasDistrict.objects.filter(
                    mas_state_id=state_id,
                    status=1  # Only active districts
                ).values('id', 'name', 'code')
                print(f"Found {districts.count()} districts for state_id={state_id}")
                return JsonResponse({'districts': list(districts)})
            except Exception as e:
                print(f"Error in get_districts_by_state: {e}")
                return JsonResponse({'error': str(e)}, status=400)
        return JsonResponse({'error': 'Missing state_id'}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=400)

@login_required
def get_district_details(request):
    """AJAX endpoint to get district details by district ID"""
    if request.method == 'GET':
        district_id = request.GET.get('district_id')
        if district_id:
            try:
                district = MasDistrict.objects.get(id=district_id, status=1)
                return JsonResponse({
                    'district_code': district.code,
                    'district_name': district.name
                })
            except MasDistrict.DoesNotExist:
                return JsonResponse({'error': 'District not found'}, status=404)
            except Exception as e:
                print(f"Error in get_district_details: {e}")
                return JsonResponse({'error': str(e)}, status=400)
        return JsonResponse({'error': 'Missing district_id'}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=400)

# Utility function to check user menu access
def check_user_menu_access(user, url_name, permission_type='can_view'):
    """Check if a user has access to a specific menu item with the specified permission"""
    if not user.is_authenticated:
        return False
    
    # Superusers have access to everything
    if user.is_superuser:
        return True
    
    # Check if user has specific menu access
    try:
        user_access = UserMenuAccess.objects.get(
            user=user,
            menu_item__url_name=url_name,
            is_active=True
        )
        
        if permission_type == 'can_view':
            return user_access.can_view
        elif permission_type == 'can_create':
            return user_access.can_create
        elif permission_type == 'can_edit':
            return user_access.can_edit
        elif permission_type == 'can_delete':
            return user_access.can_delete
        elif permission_type == 'can_approve':
            return user_access.can_approve
        
        return False
    except UserMenuAccess.DoesNotExist:
        # Check if menu item requires staff or superuser
        try:
            menu_item = MenuItem.objects.get(url_name=url_name, is_active=True)
            if menu_item.requires_superuser and not user.is_superuser:
                return False
            if menu_item.requires_staff and not user.is_staff:
                return False
            return True  # Default access if no specific restrictions
        except MenuItem.DoesNotExist:
            return True  # Menu item not found, allow access by default

@login_required
def user_profile(request):
    """User profile view"""
    user = request.user
    
    if request.method == 'POST':
        # Handle profile update
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        
        # Basic validation
        if not first_name or not last_name:
            messages.error(request, 'First name and last name are required.')
        elif not email:
            messages.error(request, 'Email is required.')
        else:
            # Update user profile
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save()
            
            messages.success(request, 'Profile updated successfully!')
            return redirect('user_profile')
    
    context = {
        'user': user,
        'page_title': 'User Profile'
    }
    return render(request, 'dashboard/user_profile.html', context)

@login_required
def user_settings(request):
    """User settings view"""
    user = request.user
    
    if request.method == 'POST':
        # Handle password change
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if current_password and new_password and confirm_password:
            # Verify current password
            if user.check_password(current_password):
                if new_password == confirm_password:
                    if len(new_password) >= 8:
                        user.set_password(new_password)
                        user.save()
                        
                        # Update session to prevent logout
                        update_session_auth_hash(request, user)
                        
                        messages.success(request, 'Password changed successfully!')
                        return redirect('user_settings')
                    else:
                        messages.error(request, 'New password must be at least 8 characters long.')
                else:
                    messages.error(request, 'New passwords do not match.')
            else:
                messages.error(request, 'Current password is incorrect.')
    
    context = {
        'user': user,
        'page_title': 'User Settings'
    }
    return render(request, 'dashboard/user_settings.html', context)

@login_required
def send_spo_reminder_mail(request):
    """Send manual mail reminder for SPO Rent"""
    if request.method == 'POST':
        try:
            record_id = request.POST.get('record_id')
            email = request.POST.get('email')
            spo_name = request.POST.get('spo_name')
            subject = request.POST.get('subject', 'SPO Rent Agreement Reminder')
            message = request.POST.get('message', '')
            
            # Get the SPO Rent record
            try:
                record = SPORent.objects.get(id=record_id)
            except SPORent.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'SPO Rent record not found'
                })
            
            # Prepare email content
            if not message:
                message = f"""
Dear {record.owner_name or 'Valued Partner'},

This is a reminder regarding your SPO Rent Agreement.

Key Details:
- SPO Name: {record.spo_name}
- SPO Code: {record.spo_code}
- Branch: {record.branch.state_branch_name if record.branch else 'N/A'}
- Expiry Date: {record.rental_to_date.strftime('%d/%m/%Y') if record.rental_to_date else 'N/A'}

Please ensure all necessary procedures are completed as required.

If you have any questions, please contact our support team.

Best regards,
Data Management System
                """.strip()
            
            # Send email
            try:
                from django.core.mail import send_mail
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=False,
                )
                
                # Log the successful email
                import logging
                logger = logging.getLogger(__name__)
                logger.info(
                    f'Manual reminder sent to {email} for SPO {record.spo_code} '
                    f'by user {request.user.username}'
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Email sent successfully to {email}'
                })
                
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error sending email to {email}: {str(e)}')
                return JsonResponse({
                    'success': False,
                    'message': f'Failed to send email: {str(e)}'
                })
                
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error in send_spo_reminder_mail: {str(e)}')
            return JsonResponse({
                'success': False,
                'message': f'System error: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method'
    })